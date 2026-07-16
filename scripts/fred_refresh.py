#!/usr/bin/env python3
"""Refresh FRED Venezuela datasets from public FRED series pages and CSV downloads."""

from __future__ import annotations

import csv
import datetime as dt
import html
import io
import json
import re
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

from openpyxl import Workbook

from ove_excel_format import format_sheet, write_key_values, write_table


ROOT = Path(__file__).resolve().parents[1]
FRED_ROOT = ROOT / "assets" / "data" / "fred"
RAW_DIR = FRED_ROOT / "raw"
CSV_DIR = FRED_ROOT / "csv"
JSON_DIR = FRED_ROOT / "json"
EXCEL_DIR = FRED_ROOT / "excel"
CATALOG_DIR = FRED_ROOT / "catalog"

TAG_URL = "https://fred.stlouisfed.org/tags/series?t=venezuela"
SERIES_URL = "https://fred.stlouisfed.org/series/{series_id}"
CSV_URL = "https://fred.stlouisfed.org/graph/fredgraph.csv?id={series_id}"
SOURCE_URL = "https://fred.stlouisfed.org/tags/series?t=venezuela"
USER_AGENT = "OVE FRED refresh/1.0 (+https://ove-venezuela.com/)"

DATA_FIELDS = [
  "País",
  "Fuente",
  "ID serie FRED",
  "Título",
  "Unidades",
  "Frecuencia",
  "Ajuste estacional",
  "Fecha",
  "Año",
  "Valor",
  "URL serie",
  "URL fuente",
]

CATALOG_FIELDS = [
  "Fuente",
  "País",
  "ID serie FRED",
  "Título",
  "Unidades",
  "Frecuencia",
  "Ajuste estacional",
  "Estado descarga",
  "Error descarga",
  "Número de registros",
  "Primer periodo",
  "Último periodo",
  "Último valor disponible",
  "URL serie",
  "Archivo CSV",
  "Archivo JSON",
  "Archivo Excel",
]


def today() -> str:
  return dt.date.today().isoformat()


def fetch_text(url: str) -> str:
  request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
  with urllib.request.urlopen(request, timeout=90) as response:
    return response.read().decode("utf-8-sig", errors="replace")


def fetch_bytes(url: str, timeout: int = 90) -> bytes:
  request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
  with urllib.request.urlopen(request, timeout=timeout) as response:
    return response.read()


def clean_text(value: str) -> str:
  value = re.sub(r"<[^>]+>", " ", value)
  value = html.unescape(value)
  return re.sub(r"\s+", " ", value).strip()


def split_attributes(value: str) -> tuple[str, str, str]:
  parts = [part.strip() for part in value.split(",")]
  if len(parts) >= 3:
    return ", ".join(parts[:-2]), parts[-2], parts[-1]
  if len(parts) == 2:
    return parts[0], parts[1], ""
  return value, "", ""


def infer_year(date_value: str) -> int | None:
  match = re.search(r"\d{4}", date_value or "")
  return int(match.group(0)) if match else None


def safe_float(value: str | None) -> float | None:
  if value in (None, "", "."):
    return None
  try:
    return float(value)
  except ValueError:
    return None


def parse_series_from_page(html_text: str) -> list[dict]:
  blocks = re.findall(
    r'<tr class="series-pager-title".*?(?=<tr class="series-pager-title"|</table>)',
    html_text,
    flags=re.DOTALL,
  )
  series = []
  for block in blocks:
    id_match = re.search(r'href="/series/([A-Za-z0-9_]+)"', block)
    title_match = re.search(r'class="series-title[^"]*">(.*?)</a>', block, flags=re.DOTALL)
    attr_match = re.search(r'<span class="attributes"[^>]*>(.*?)</span>', block, flags=re.DOTALL)
    dates_match = re.search(r'<span class="series-meta-dates">(.*?)</span>', block, flags=re.DOTALL)
    if not id_match or not title_match:
      continue
    series_id = id_match.group(1)
    units, frequency, seasonal = split_attributes(clean_text(attr_match.group(1)) if attr_match else "")
    series.append({
      "ID serie FRED": series_id,
      "Título": clean_text(title_match.group(1)),
      "Unidades": units,
      "Frecuencia": frequency,
      "Ajuste estacional": seasonal,
      "Rango FRED": clean_text(dates_match.group(1)) if dates_match else "",
      "URL serie": SERIES_URL.format(series_id=series_id),
    })
  return series


def load_catalog_from_fred() -> list[dict]:
  discovered: dict[str, dict] = {}
  page = 1
  while True:
    url = f"{TAG_URL}&{urllib.parse.urlencode({'pageID': page})}"
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = RAW_DIR / f"fred_venezuela_tag_page_{page}.html"
    try:
      html_text = fetch_text(url)
      cache_path.write_text(html_text, encoding="utf-8")
    except urllib.error.HTTPError:
      if not cache_path.exists():
        raise
      html_text = cache_path.read_text(encoding="utf-8")
    page_series = parse_series_from_page(html_text)
    for item in page_series:
      discovered[item["ID serie FRED"]] = item
    if f'pageID={page + 1}' not in html_text:
      break
    page += 1
    if page > 25:
      raise RuntimeError("FRED pagination exceeded expected limit")
  return list(discovered.values())


def observation_row(series: dict, date_value: str, value: float) -> dict:
  return {
      "País": "Venezuela",
      "Fuente": "FRED - Federal Reserve Bank of St. Louis",
      "ID serie FRED": series["ID serie FRED"],
      "Título": series["Título"],
      "Unidades": series["Unidades"],
      "Frecuencia": series["Frecuencia"],
      "Ajuste estacional": series["Ajuste estacional"],
      "Fecha": date_value,
      "Año": infer_year(date_value),
      "Valor": value,
      "URL serie": series["URL serie"],
      "URL fuente": SOURCE_URL,
    }


def parse_observation_csv(text: str, series_map: dict[str, dict]) -> list[dict]:
  rows = list(csv.DictReader(io.StringIO(text)))
  observations = []
  for row in rows:
    date_value = row.get("observation_date", "")
    if not date_value:
      continue
    for series_id, series in series_map.items():
      value = safe_float(row.get(series_id))
      if value is not None:
        observations.append(observation_row(series, date_value, value))
  return observations


def download_batch_observations(batch: list[dict]) -> tuple[list[dict], list[dict]]:
  series_map = {series["ID serie FRED"]: series for series in batch}
  url = CSV_URL.format(series_id=",".join(series_map))
  errors = []
  try:
    content = fetch_bytes(url, timeout=120)
  except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
    return [], [{
      "ID serie FRED": ",".join(series_map),
      "Título": "Bloque FRED",
      "Error": str(exc),
    }]

  observations = []
  if content.startswith(b"PK"):
    with zipfile.ZipFile(io.BytesIO(content)) as archive:
      for name in archive.namelist():
        if not name.endswith(".csv"):
          continue
        observations.extend(parse_observation_csv(archive.read(name).decode("utf-8-sig", errors="replace"), series_map))
  else:
    observations.extend(parse_observation_csv(content.decode("utf-8-sig", errors="replace"), series_map))

  seen = {row["ID serie FRED"] for row in observations}
  for series_id, series in series_map.items():
    if series_id not in seen:
      errors.append({
        "ID serie FRED": series_id,
        "Título": series["Título"],
        "Error": "Sin observaciones descargadas desde FRED",
      })
  return observations, errors


def download_single_observations(series: dict) -> tuple[list[dict], list[dict]]:
  series_id = series["ID serie FRED"]
  try:
    content = fetch_bytes(CSV_URL.format(series_id=series_id), timeout=30)
    observations = parse_observation_csv(content.decode("utf-8-sig", errors="replace"), {series_id: series})
  except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError, csv.Error) as exc:
    return [], [{
      "ID serie FRED": series_id,
      "Título": series["Título"],
      "Error": str(exc),
    }]
  if not observations:
    return [], [{
      "ID serie FRED": series_id,
      "Título": series["Título"],
      "Error": "Sin observaciones descargadas desde FRED",
    }]
  return observations, []


def load_data(catalog: list[dict]) -> tuple[list[dict], list[dict]]:
  all_rows = []
  download_errors = []
  with ThreadPoolExecutor(max_workers=8) as executor:
    futures = [executor.submit(download_single_observations, series) for series in catalog]
    for future in as_completed(futures):
      rows, errors = future.result()
      all_rows.extend(rows)
      download_errors.extend(errors)
  all_rows.sort(key=lambda item: (item["ID serie FRED"], item["Fecha"]))
  return all_rows, download_errors


def write_data_outputs(rows: list[dict], generated_at: str, errors: list[dict]) -> None:
  for directory in (CSV_DIR, JSON_DIR, EXCEL_DIR, CATALOG_DIR):
    directory.mkdir(parents=True, exist_ok=True)
  csv_path = CSV_DIR / "ove_fred_venezuela.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=DATA_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)

  payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "FRED - Venezuela",
      "Fecha generación": generated_at,
      "Fuente": "FRED - Federal Reserve Bank of St. Louis",
      "URL fuente": SOURCE_URL,
      "País": "Venezuela",
      "Número de indicadores": len({row["ID serie FRED"] for row in rows}),
      "Número de registros": len(rows),
      "Errores descarga": len(errors),
    },
    "errores": errors,
    "datos": rows,
  }
  (JSON_DIR / "ove_fred_venezuela.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb = Workbook()
  ws = wb.active
  ws.title = "datos"
  format_sheet(ws, "FRED - Venezuela", "Series públicas etiquetadas para Venezuela")
  write_table(ws, DATA_FIELDS, rows)
  meta = wb.create_sheet("metadatos")
  format_sheet(meta, "Metadatos", "FRED - Venezuela")
  write_key_values(meta, payload["metadatos"].items())
  if errors:
    error_ws = wb.create_sheet("errores")
    format_sheet(error_ws, "Errores de descarga", "FRED - Venezuela")
    write_table(error_ws, ["ID serie FRED", "Título", "Error"], errors)
  wb.save(EXCEL_DIR / "ove_fred_venezuela.xlsx")


def catalog_rows(rows: list[dict], discovered: list[dict], errors: list[dict]) -> list[dict]:
  grouped: dict[str, list[dict]] = defaultdict(list)
  for row in rows:
    grouped[row["ID serie FRED"]].append(row)
  metadata = {item["ID serie FRED"]: item for item in discovered}
  error_map = {item["ID serie FRED"]: item["Error"] for item in errors}
  catalog = []
  for series_id, series in sorted(metadata.items()):
    items = grouped.get(series_id, [])
    latest = max(items, key=lambda item: item["Fecha"]) if items else None
    catalog_row = {
      "Fuente": "FRED - Federal Reserve Bank of St. Louis",
      "País": "Venezuela",
      "ID serie FRED": series_id,
      "Título": series["Título"],
      "Unidades": series["Unidades"],
      "Frecuencia": series["Frecuencia"],
      "Ajuste estacional": series["Ajuste estacional"],
      "Estado descarga": "Con datos" if items else "No descargada",
      "Error descarga": error_map.get(series_id, ""),
      "Número de registros": len(items),
      "Primer periodo": min((item["Fecha"] for item in items), default=""),
      "Último periodo": latest["Fecha"] if latest else "",
      "Último valor disponible": latest["Valor"] if latest else None,
      "URL serie": series["URL serie"],
      "Archivo CSV": "assets/data/fred/csv/ove_fred_venezuela.csv",
      "Archivo JSON": "assets/data/fred/json/ove_fred_venezuela.json",
      "Archivo Excel": "assets/data/fred/excel/ove_fred_venezuela.xlsx",
    }
    catalog.append(catalog_row)
  return catalog


def write_catalog(catalog: list[dict], generated_at: str) -> None:
  csv_path = CATALOG_DIR / "catalogo_dataset_web_ove_fred.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=CATALOG_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(catalog)
  payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "FRED - Venezuela",
      "Fecha generación": generated_at,
      "Fuente": "FRED - Federal Reserve Bank of St. Louis",
      "URL fuente": SOURCE_URL,
      "País": "Venezuela",
      "Número de series catalogadas": len(catalog),
      "Número de series con datos": sum(1 for row in catalog if row["Número de registros"] > 0),
      "Número de registros": sum(row["Número de registros"] for row in catalog),
    },
    "catalogo": catalog,
  }
  (CATALOG_DIR / "fred-catalog.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb = Workbook()
  ws = wb.active
  ws.title = "catalogo"
  format_sheet(ws, "FRED - Venezuela", "Catálogo de series")
  write_table(ws, CATALOG_FIELDS, catalog)
  meta = wb.create_sheet("metadatos")
  format_sheet(meta, "Metadatos", "Catálogo FRED")
  write_key_values(meta, payload["metadatos"].items())
  wb.save(CATALOG_DIR / "catalogo_dataset_web_ove_fred.xlsx")


def main() -> int:
  generated_at = today()
  catalog = load_catalog_from_fred()
  rows, errors = load_data(catalog)
  if not rows and (JSON_DIR / "ove_fred_venezuela.json").exists():
    existing = json.loads((JSON_DIR / "ove_fred_venezuela.json").read_text(encoding="utf-8"))
    existing_rows = existing.get("datos", [])
    if existing_rows:
      print(json.dumps({
        "dataset": "FRED - Venezuela",
        "generated_at": generated_at,
        "series_discovered": len(catalog),
        "series_with_data": existing.get("metadatos", {}).get("Número de indicadores", 0),
        "records": len(existing_rows),
        "download_errors": len(errors),
        "preserved_existing_data": True,
        "output": str(FRED_ROOT),
      }, ensure_ascii=False))
      return 0
  write_data_outputs(rows, generated_at, errors)
  structured_catalog = catalog_rows(rows, catalog, errors)
  write_catalog(structured_catalog, generated_at)
  print(json.dumps({
    "dataset": "FRED - Venezuela",
    "generated_at": generated_at,
    "series_discovered": len(catalog),
    "series_with_data": len(structured_catalog),
    "records": len(rows),
    "download_errors": len(errors),
    "output": str(FRED_ROOT),
  }, ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
