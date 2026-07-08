#!/usr/bin/env python3
"""Refresh OVE World Bank Venezuela datasets from the public API."""

from __future__ import annotations

import csv
import datetime as dt
import json
import math
import time
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
WB_ROOT = ROOT / "assets" / "data" / "world-bank"
JSON_DIR = WB_ROOT / "json"
CSV_DIR = WB_ROOT / "csv"
EXCEL_DIR = WB_ROOT / "excel"
CATALOG_DIR = WB_ROOT / "catalog"
CATALOG_PATH = CATALOG_DIR / "world-bank-catalog.json"

FIELDNAMES = [
  "País OVE",
  "País Banco Mundial",
  "Código ISO 2",
  "Código ISO 3",
  "Región",
  "Nivel de ingreso",
  "Año",
  "Área temática",
  "ID área",
  "Subárea",
  "ID subárea",
  "Código indicador",
  "Indicador",
  "Valor",
  "Fuente",
  "Fecha descarga",
  "Error descarga",
]


def today() -> str:
  return dt.date.today().isoformat()


def load_json(path: Path):
  return json.loads(path.read_text(encoding="utf-8-sig"))


def fetch_indicator(code: str) -> tuple[list[dict], str | None]:
  params = urllib.parse.urlencode({
    "format": "json",
    "per_page": 20000,
    "date": f"1960:{dt.date.today().year + 1}",
  })
  url = f"https://api.worldbank.org/v2/country/VEN/indicator/{code}?{params}"
  request = urllib.request.Request(url, headers={"User-Agent": "OVE data refresh/1.0"})
  try:
    with urllib.request.urlopen(request, timeout=45) as response:
      payload = json.loads(response.read().decode("utf-8"))
  except Exception as exc:
    return [], str(exc)
  if not isinstance(payload, list) or len(payload) < 2:
    return [], "Unexpected World Bank API response"
  return payload[1] or [], None


def indicator_plan() -> dict[str, dict]:
  plan = {}
  for path in sorted(JSON_DIR.glob("ove_banco_mundial_venezuela_*.json")):
    data = load_json(path)
    for row in data.get("datos", []):
      code = row["Código indicador"]
      plan.setdefault(code, {
        "Área temática": row["Área temática"],
        "ID área": row["ID área"],
        "Subárea": row["Subárea"],
        "ID subárea": row["ID subárea"],
        "Indicador": row["Indicador"],
      })
  return plan


def safe_float(value):
  if value is None or value == "":
    return None
  try:
    number = float(value)
  except (TypeError, ValueError):
    return None
  return None if math.isnan(number) else number


def row_from_observation(obs: dict, meta: dict, code: str, fetched: str, error: str | None = None) -> dict:
  country = obs.get("country") or {}
  return {
    "País OVE": "Venezuela",
    "País Banco Mundial": country.get("value") or "Venezuela, RB",
    "Código ISO 2": "VE",
    "Código ISO 3": "VEN",
    "Región": "Latin America & Caribbean",
    "Nivel de ingreso": "Not classified",
    "Año": int(obs["date"]),
    "Área temática": meta["Área temática"],
    "ID área": meta["ID área"],
    "Subárea": meta["Subárea"],
    "ID subárea": meta["ID subárea"],
    "Código indicador": code,
    "Indicador": meta["Indicador"],
    "Valor": safe_float(obs.get("value")),
    "Fuente": "Banco Mundial - World Development Indicators",
    "Fecha descarga": fetched,
    "Error descarga": error,
  }


def write_area_files(area_id: str, area_name: str, rows: list[dict], fetched: str) -> dict:
  rows.sort(key=lambda item: (item["Código indicador"], item["Año"]))
  slug = f"ove_banco_mundial_venezuela_{area_id}"
  csv_path = CSV_DIR / f"{slug}.csv"
  json_path = JSON_DIR / f"{slug}.json"
  xlsx_path = EXCEL_DIR / f"{slug}.xlsx"
  for path in (csv_path, json_path, xlsx_path):
    path.parent.mkdir(parents=True, exist_ok=True)

  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=FIELDNAMES, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)

  payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": f"Banco Mundial - Venezuela - {area_name}",
      "Fecha generación": fetched,
      "Número de indicadores": len({row["Código indicador"] for row in rows}),
      "Número de registros": len(rows),
    },
    "datos": rows,
  }
  json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb = Workbook()
  ws = wb.active
  ws.title = "datos"
  ws.append(FIELDNAMES)
  for row in rows:
    ws.append([row.get(field) for field in FIELDNAMES])
  meta = wb.create_sheet("metadatos")
  for key, value in payload["metadatos"].items():
    meta.append([key, value])
  wb.save(xlsx_path)

  years = [row["Año"] for row in rows]
  return {
    "Área temática": area_name,
    "ID área": area_id,
    "Registros": len(rows),
    "Indicadores": len({row["Código indicador"] for row in rows}),
    "Primer año": min(years) if years else None,
    "Último año": max(years) if years else None,
    "Último año con dato": max((row["Año"] for row in rows if row["Valor"] is not None), default=None),
    "Archivo CSV": f"assets/data/world-bank/csv/{slug}.csv",
    "Archivo JSON": f"assets/data/world-bank/json/{slug}.json",
    "Archivo Excel": f"assets/data/world-bank/excel/{slug}.xlsx",
  }


def latest_non_null(rows: list[dict], code: str) -> dict | None:
  candidates = [row for row in rows if row["Código indicador"] == code and row["Valor"] is not None]
  return max(candidates, key=lambda item: item["Año"]) if candidates else None


def build_latest_summary(rows: list[dict], fetched: str) -> None:
  selected = {
    "population_total": "SP.POP.TOTL",
    "gdp_current_usd": "NY.GDP.MKTP.CD",
    "gdp_growth_real": "NY.GDP.MKTP.KD.ZG",
    "gdp_per_capita_current_usd": "NY.GDP.PCAP.CD",
    "inflation_cpi": "FP.CPI.TOTL.ZG",
    "unemployment_total": "SL.UEM.TOTL.ZS",
    "labor_force_total": "SL.TLF.TOTL.IN",
    "exports_goods_services_usd": "NE.EXP.GNFS.CD",
    "imports_goods_services_usd": "NE.IMP.GNFS.CD",
    "current_account_usd": "BN.CAB.XOKA.CD",
    "women_parliament": "SG.GEN.PARL.ZS",
    "internet_users": "IT.NET.USER.ZS",
  }
  indicators = {}
  for key, code in selected.items():
    row = latest_non_null(rows, code)
    if row:
      indicators[key] = {
        "code": code,
        "indicator": row["Indicador"],
        "value": row["Valor"],
        "year": row["Año"],
        "source": row["Fuente"],
      }
  payload = {
    "metadata": {
      "source": "Banco Mundial - World Development Indicators",
      "country": "Venezuela, RB",
      "last_fetched_at": fetched,
      "status": "official_source",
    },
    "indicators": indicators,
  }
  (CATALOG_DIR / "world-bank-latest-summary.json").write_text(
    json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
  )


def write_catalog_tables(catalog: list[dict]) -> None:
  if not catalog:
    return
  fields = list(catalog[0].keys())
  csv_path = CATALOG_DIR / "catalogo_dataset_web_ove_banco_mundial.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=fields, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(catalog)

  wb = Workbook()
  ws = wb.active
  ws.title = "catalogo"
  ws.append(fields)
  for row in catalog:
    ws.append([row.get(field) for field in fields])
  wb.save(CATALOG_DIR / "catalogo_dataset_web_ove_banco_mundial.xlsx")


def main() -> int:
  fetched = today()
  plan = indicator_plan()
  rows_by_area: dict[str, list[dict]] = defaultdict(list)
  all_rows: list[dict] = []
  for index, (code, meta) in enumerate(sorted(plan.items()), start=1):
    observations, error = fetch_indicator(code)
    if error:
      years = [{"date": str(year), "value": None, "country": {"value": "Venezuela, RB"}} for year in range(1960, dt.date.today().year + 1)]
      observations = years
    for obs in observations:
      if not str(obs.get("date", "")).isdigit():
        continue
      row = row_from_observation(obs, meta, code, fetched, error)
      rows_by_area[meta["ID área"]].append(row)
      all_rows.append(row)
    if index % 10 == 0:
      time.sleep(0.5)

  catalog = []
  for area_id, rows in sorted(rows_by_area.items()):
    area_name = rows[0]["Área temática"]
    catalog.append(write_area_files(area_id, area_name, rows, fetched))

  catalog_payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "Banco Mundial - Venezuela",
      "Fecha generación": fetched,
      "Número de áreas": len(catalog),
      "Número de indicadores": len(plan),
      "Número de registros": len(all_rows),
    },
    "catalogo": catalog,
  }
  CATALOG_PATH.write_text(json.dumps(catalog_payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
  write_catalog_tables(catalog)
  build_latest_summary(all_rows, fetched)
  print(json.dumps(catalog_payload["metadatos"], ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
