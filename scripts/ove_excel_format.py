"""OVE corporate formatting helpers for public Excel workbooks."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Sequence
import datetime as dt

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
  from openpyxl.drawing.image import Image
except ImportError:  # pragma: no cover - optional Pillow dependency guard.
  Image = None


ROOT = Path(__file__).resolve().parents[1]
LOGO_PATH = ROOT / "assets" / "ove-logo-brand.png"
TABLE_START_ROW = 6
NAVY = "0A2D5A"
YELLOW = "FFC20E"
WHITE = "FFFFFF"
GRAY = "6B6B6B"
SPANISH_INTEGER_FORMAT = '[$-es-ES]#,##0'
SPANISH_DECIMAL_FORMAT = '[$-es-ES]#,##0.00'
SPANISH_DATE_FORMAT = '[$-es-ES]dd/mm/yyyy'
SPANISH_DATETIME_FORMAT = '[$-es-ES]dd/mm/yyyy hh:mm'
SPANISH_HEADER_LABELS = {
  "indicator_id": "ID indicador",
  "indicator_name": "Indicador",
  "indicator": "Indicador",
  "area": "Área",
  "category": "Categoría",
  "date": "Fecha",
  "year": "Año",
  "month": "Mes",
  "month_name": "Nombre del mes",
  "period": "Periodo",
  "value": "Valor",
  "value_buy": "Valor compra",
  "value_sell": "Valor venta",
  "index_value": "Valor del índice",
  "monthly_variation_pct": "Variación mensual (%)",
  "annual_real_gdp_growth_pct": "Crecimiento anual PIB real (%)",
  "currency": "Moneda",
  "unit": "Unidad",
  "frequency": "Frecuencia",
  "source": "Fuente",
  "source_url": "URL fuente",
  "fetched_at": "Fecha de captura",
}


def add_ove_header(ws, title: str | None = None, subtitle: str | None = None) -> None:
  for row in range(1, TABLE_START_ROW):
    ws.row_dimensions[row].height = 20
  ws.row_dimensions[1].height = 28
  ws.row_dimensions[2].height = 22
  ws.row_dimensions[3].height = 18
  ws.row_dimensions[4].height = 16
  ws.row_dimensions[5].height = 8

  if Image is not None and LOGO_PATH.exists():
    logo = Image(str(LOGO_PATH))
    logo.width = 190
    logo.height = 48
    ws.add_image(logo, "A1")
  else:
    ws["A1"] = "OVE"
    ws["A1"].font = Font(bold=True, color=NAVY, size=18)

  if title:
    ws["D1"] = title
    ws["D1"].font = Font(bold=True, color=NAVY, size=14)
  if subtitle:
    ws["D2"] = subtitle
    ws["D2"].font = Font(color=GRAY, size=10)

  ws.sheet_view.showGridLines = False


def write_table(ws, fields: Sequence[str], rows: Iterable[dict], start_row: int = TABLE_START_ROW) -> None:
  for col_idx, field in enumerate(fields, start=1):
    cell = ws.cell(row=start_row, column=col_idx, value=SPANISH_HEADER_LABELS.get(field, field))
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws.row_dimensions[start_row].height = 24

  for row_idx, row in enumerate(rows, start=start_row + 1):
    for col_idx, field in enumerate(fields, start=1):
      cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field))
      cell.alignment = Alignment(vertical="top", wrap_text=True)
      apply_spanish_number_format(cell)

  ws.freeze_panes = f"A{start_row + 1}"
  apply_widths(ws, len(fields))


def write_key_values(ws, items: Iterable[tuple[str, object]], start_row: int = TABLE_START_ROW) -> None:
  ws.cell(row=start_row, column=1, value="Campo")
  ws.cell(row=start_row, column=2, value="Valor")
  for col_idx in (1, 2):
    cell = ws.cell(row=start_row, column=col_idx)
    cell.font = Font(bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="center", vertical="center")
  for row_idx, (key, value) in enumerate(items, start=start_row + 1):
    ws.cell(row=row_idx, column=1, value=key)
    ws.cell(row=row_idx, column=2, value=value)
    ws.cell(row=row_idx, column=1).font = Font(bold=True, color=NAVY)
    ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    apply_spanish_number_format(ws.cell(row=row_idx, column=2))
  ws.freeze_panes = f"A{start_row + 1}"
  apply_widths(ws, 2)


def apply_widths(ws, column_count: int) -> None:
  for col_idx in range(1, column_count + 1):
    width = 24 if col_idx <= 2 else 18
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def format_sheet(ws, title: str | None = None, subtitle: str | None = None) -> None:
  add_ove_header(ws, title=title, subtitle=subtitle)


def apply_spanish_number_format(cell) -> None:
  value = cell.value
  if isinstance(value, bool) or value is None:
    return
  if isinstance(value, dt.datetime):
    cell.number_format = SPANISH_DATETIME_FORMAT
    return
  if isinstance(value, dt.date):
    cell.number_format = SPANISH_DATE_FORMAT
    return
  if isinstance(value, int):
    cell.number_format = SPANISH_INTEGER_FORMAT
    return
  if isinstance(value, float):
    cell.number_format = SPANISH_INTEGER_FORMAT if value.is_integer() else SPANISH_DECIMAL_FORMAT
