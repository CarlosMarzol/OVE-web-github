#!/usr/bin/env python3
"""Refresh ILOSTAT/OIT Venezuela datasets and catalog them for OVE."""

from __future__ import annotations

import csv
import datetime as dt
import gzip
import io
import json
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook


ROOT = Path(__file__).resolve().parents[1]
ILO_ROOT = ROOT / "assets" / "data" / "ilo"
RAW_DIR = ILO_ROOT / "raw"
CSV_DIR = ILO_ROOT / "csv"
JSON_DIR = ILO_ROOT / "json"
EXCEL_DIR = ILO_ROOT / "excel"
CATALOG_DIR = ILO_ROOT / "catalog"

FREQUENCIES = {
  "A": {"label": "Anual", "slug": "anual"},
  "Q": {"label": "Trimestral", "slug": "trimestral"},
  "M": {"label": "Mensual", "slug": "mensual"},
}

CATALOG_FIELDS = [
  "Fuente",
  "País",
  "Frecuencia",
  "Código indicador",
  "Indicador",
  "Tema",
  "Base de datos",
  "Dimensiones disponibles",
  "Número de series",
  "Número de registros",
  "Primer periodo",
  "Último periodo",
  "Último periodo con dato",
  "Último valor disponible",
  "Última actualización OIT",
  "Archivo datos comprimidos",
]


def today() -> str:
  return dt.date.today().isoformat()


def fetch_bytes(url: str) -> bytes:
  request = urllib.request.Request(url, headers={"User-Agent": "OVE ILOSTAT refresh/1.0"})
  with urllib.request.urlopen(request, timeout=120) as response:
    return response.read()


def write_bytes(path: Path, content: bytes) -> None:
  path.parent.mkdir(parents=True, exist_ok=True)
  path.write_bytes(content)


def read_csv_bytes(content: bytes) -> list[dict]:
  text = content.decode("utf-8-sig")
  return list(csv.DictReader(io.StringIO(text)))


def load_metadata() -> dict[tuple[str, str], dict]:
  content = fetch_bytes("https://rplumber.ilo.org/metadata/toc/indicator?lang=es&format=.csv")
  write_bytes(RAW_DIR / "ilostat_indicator_toc_es.csv", content)
  metadata = {}
  for row in read_csv_bytes(content):
    metadata[(row.get("indicator", ""), row.get("freq", ""))] = row
  return metadata


def period_key(value: str) -> tuple[int, int, int]:
  if not value:
    return (0, 0, 0)
  if "Q" in value:
    year, quarter = value.split("Q", 1)
    return (int(year), int(quarter), 0)
  if "M" in value:
    year, month = value.split("M", 1)
    return (int(year), 0, int(month))
  return (int(value), 0, 0)


def safe_float(value: str | None) -> float | None:
  if value in (None, ""):
    return None
  try:
    return float(value)
  except ValueError:
    return None


def iter_rows_from_gzip(path: Path):
  with gzip.open(path, "rt", encoding="utf-8-sig", newline="") as handle:
    yield from csv.DictReader(handle)


def download_frequency(freq: str) -> dict:
  info = FREQUENCIES[freq]
  url = f"https://rplumber.ilo.org/data/ref_area?format=.csv.gz&id=VEN_{freq}"
  raw_name = f"ove_oit_ilostat_venezuela_{info['slug']}.csv.gz"
  raw_path = CSV_DIR / raw_name
  content = fetch_bytes(url)
  write_bytes(raw_path, content)
  return {
    "frequency": freq,
    "label": info["label"],
    "slug": info["slug"],
    "url": url,
    "raw_path": raw_path,
    "asset_path": f"assets/data/ilo/csv/{raw_name}",
    "bytes": len(content),
  }


def summarize(downloads: list[dict], metadata: dict[tuple[str, str], dict]) -> tuple[list[dict], list[dict]]:
  catalog = []
  frequency_summary = []
  for download in downloads:
    freq = download["frequency"]
    groups: dict[str, dict] = {}
    periods = []
    record_count = 0
    for row in iter_rows_from_gzip(download["raw_path"]):
      record_count += 1
      indicator = row["indicator"]
      period = row["time"]
      periods.append(period)
      item = groups.setdefault(indicator, {
        "records": 0,
        "series": set(),
        "periods": [],
        "latest_period": None,
        "latest_value": None,
      })
      item["records"] += 1
      item["series"].add("|".join([
        row.get("sex", ""),
        row.get("classif1", ""),
        row.get("classif2", ""),
        row.get("source", ""),
      ]))
      item["periods"].append(period)
      value = safe_float(row.get("obs_value"))
      if value is not None and (item["latest_period"] is None or period_key(period) > period_key(item["latest_period"])):
        item["latest_period"] = period
        item["latest_value"] = value

    for indicator, stats in sorted(groups.items()):
      meta = metadata.get((indicator, freq), {})
      dimensions = meta.get("classif.labels") or ", ".join(part for part in [meta.get("rep_var.label"), meta.get("classification")] if part)
      catalog.append({
        "Fuente": "OIT - ILOSTAT",
        "País": "Venezuela",
        "Frecuencia": download["label"],
        "Código indicador": indicator,
        "Indicador": meta.get("indicator.label") or indicator,
        "Tema": meta.get("subject.label") or "",
        "Base de datos": meta.get("database.label") or "",
        "Dimensiones disponibles": dimensions,
        "Número de series": len(stats["series"]),
        "Número de registros": stats["records"],
        "Primer periodo": min(stats["periods"], key=period_key) if stats["periods"] else None,
        "Último periodo": max(stats["periods"], key=period_key) if stats["periods"] else None,
        "Último periodo con dato": stats["latest_period"],
        "Último valor disponible": stats["latest_value"],
        "Última actualización OIT": meta.get("last.update") or "",
        "Archivo datos comprimidos": download["asset_path"],
      })
    frequency_summary.append({
      "Frecuencia": download["label"],
      "Código frecuencia": freq,
      "Indicadores": len(groups),
      "Registros": record_count,
      "Primer periodo": min(periods, key=period_key) if periods else None,
      "Último periodo": max(periods, key=period_key) if periods else None,
      "Archivo datos comprimidos": download["asset_path"],
      "Fuente URL": download["url"],
      "Tamaño bytes": download["bytes"],
    })
  return catalog, frequency_summary


def write_table_outputs(catalog: list[dict], frequency_summary: list[dict], generated_at: str) -> None:
  for directory in (CSV_DIR, JSON_DIR, EXCEL_DIR, CATALOG_DIR):
    directory.mkdir(parents=True, exist_ok=True)

  csv_path = CATALOG_DIR / "catalogo_dataset_web_ove_oit_ilostat.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=CATALOG_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(catalog)

  catalog_json = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "OIT - ILOSTAT - Venezuela",
      "Fecha generación": generated_at,
      "Fuente": "OIT - ILOSTAT bulk API",
      "País": "Venezuela",
      "Número de frecuencias": len(frequency_summary),
      "Número de indicadores": len({row["Código indicador"] for row in catalog}),
      "Número de registros": sum(row["Registros"] for row in frequency_summary),
    },
    "frecuencias": frequency_summary,
    "catalogo": catalog,
  }
  (CATALOG_DIR / "ilo-catalog.json").write_text(json.dumps(catalog_json, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
  (JSON_DIR / "ove_oit_ilostat_venezuela_catalogo_series.json").write_text(
    json.dumps(catalog_json, ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
  )

  series_csv = CSV_DIR / "ove_oit_ilostat_venezuela_catalogo_series.csv"
  with series_csv.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=CATALOG_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(catalog)

  wb = Workbook()
  ws = wb.active
  ws.title = "catalogo"
  ws.append(CATALOG_FIELDS)
  for row in catalog:
    ws.append([row.get(field) for field in CATALOG_FIELDS])
  summary = wb.create_sheet("frecuencias")
  summary_fields = list(frequency_summary[0].keys()) if frequency_summary else []
  summary.append(summary_fields)
  for row in frequency_summary:
    summary.append([row.get(field) for field in summary_fields])
  meta = wb.create_sheet("metadatos")
  for key, value in catalog_json["metadatos"].items():
    meta.append([key, value])
  wb.save(EXCEL_DIR / "ove_oit_ilostat_venezuela_catalogo_series.xlsx")
  wb.save(CATALOG_DIR / "catalogo_dataset_web_ove_oit_ilostat.xlsx")


def main() -> int:
  generated_at = today()
  metadata = load_metadata()
  downloads = [download_frequency(freq) for freq in FREQUENCIES]
  catalog, frequency_summary = summarize(downloads, metadata)
  write_table_outputs(catalog, frequency_summary, generated_at)
  print(json.dumps({
    "dataset": "OIT - ILOSTAT - Venezuela",
    "generated_at": generated_at,
    "frequencies": len(frequency_summary),
    "indicators": len({row["Código indicador"] for row in catalog}),
    "records": sum(row["Registros"] for row in frequency_summary),
    "output": str(ILO_ROOT),
  }, ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
