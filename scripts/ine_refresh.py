#!/usr/bin/env python3
"""Refresh INE Venezuela resource catalog from the public INE website."""

from __future__ import annotations

import csv
import datetime as dt
import gzip
import html
import json
import re
import sys
import tempfile
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ove_excel_format import format_sheet, write_key_values, write_table

try:
  import xlrd
except ImportError:  # pragma: no cover - workflow installs xlrd.
  xlrd = None


ROOT = Path(__file__).resolve().parents[1]
INE_ROOT = ROOT / "assets" / "data" / "ine"
RAW_DIR = INE_ROOT / "raw"
CATALOG_DIR = INE_ROOT / "catalog"
CSV_DIR = INE_ROOT / "csv"
JSON_DIR = INE_ROOT / "json"
EXCEL_DIR = INE_ROOT / "excel"

HOME_URL = "https://ine.gob.ve/"
SOURCE_NAME = "INE Venezuela - Instituto Nacional de Estadística"
USER_AGENT = "OVE INE refresh/1.0 (+https://ove-web-github.vercel.app/)"

CATALOG_FIELDS = [
  "Fuente",
  "País",
  "Título",
  "Categoría",
  "Tipo de recurso",
  "Formato",
  "Año publicación",
  "Mes publicación",
  "Archivo original",
  "URL fuente",
  "URL portal",
]

CELL_FIELDS = [
  "ID recurso",
  "Título",
  "Categoría",
  "Formato",
  "URL fuente",
  "Hoja",
  "Fila",
  "Columna",
  "Celda",
  "Valor",
  "Tipo valor",
]

SHEET_FIELDS = [
  "ID recurso",
  "Título",
  "Categoría",
  "Formato",
  "URL fuente",
  "Hoja",
  "Filas detectadas",
  "Columnas detectadas",
  "Celdas con valor",
  "Estado extracción",
  "Error",
]

MONTHS = {
  "01": "enero",
  "02": "febrero",
  "03": "marzo",
  "04": "abril",
  "05": "mayo",
  "06": "junio",
  "07": "julio",
  "08": "agosto",
  "09": "septiembre",
  "10": "octubre",
  "11": "noviembre",
  "12": "diciembre",
}


def today() -> str:
  return dt.date.today().isoformat()


def fetch_text(url: str) -> str:
  request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
  with urllib.request.urlopen(request, timeout=90) as response:
    return response.read().decode("utf-8", errors="replace")


def fetch_bytes(url: str) -> bytes:
  request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
  with urllib.request.urlopen(request, timeout=35) as response:
    return response.read()


def clean_text(value: str) -> str:
  value = re.sub(r"<[^>]+>", " ", value)
  value = html.unescape(value)
  return re.sub(r"\s+", " ", value).strip()


def extension(url: str) -> str:
  path = urllib.parse.urlparse(url).path
  if "." not in path:
    return ""
  return path.rsplit(".", 1)[-1].lower()


def publication_date_parts(url: str) -> tuple[str, str]:
  match = re.search(r"/uploads/(\d{4})/(\d{2})/", url)
  if not match:
    return "", ""
  year, month = match.groups()
  return year, MONTHS.get(month, month)


def infer_category(title: str, url: str) -> str:
  text = f"{title} {url}".lower()
  checks = [
    ("Censo 2011", ["censo-2011", "censo_2011", "censo 2011"]),
    ("Censo 2001-2011", ["2001-y-2011", "2001-2011"]),
    ("Proyección de población", ["proyeccion", "proyección", "edad-simple", "esperanza-de-vida", "nacional-36"]),
    ("Población", ["poblacion", "población", "grupos-edad", "sexo"]),
    ("Estadísticas vitales", ["nacimientos", "defunciones", "matrimonios", "suicidios", "divorcios"]),
    ("Vivienda y servicios", ["vivienda", "viviendas", "hogares", "servicio", "agua", "basura", "electrico", "excretas", "hacinamiento"]),
    ("Comercio exterior", ["exportaciones", "importaciones"]),
    ("Encuesta de hogares", ["por-tipo", "por-variable", "trabajo-2001-2023", "encuesta-de-hogares"]),
    ("Nomenclador", ["nomenclador"]),
    ("Anuarios", ["anuario"]),
    ("Publicaciones", ["resumen", "folleto", "plan", "revista", "documento", "pdf"]),
  ]
  for category, tokens in checks:
    if any(token in text for token in tokens):
      return category
  return "Otros recursos INE"


def resource_type(fmt: str) -> str:
  if fmt in {"xls", "xlsx", "csv"}:
    return "Dato tabular"
  if fmt == "pdf":
    return "Documento"
  return "Recurso"


def resource_id(row: dict) -> str:
  stem = Path(urllib.parse.urlparse(row["URL fuente"]).path).stem.lower()
  stem = re.sub(r"[^a-z0-9]+", "_", stem).strip("_")
  return stem[:90] or "recurso_ine"


def value_type(value) -> str:
  if value is None:
    return "vacío"
  if isinstance(value, bool):
    return "booleano"
  if isinstance(value, (int, float)):
    return "numérico"
  if isinstance(value, dt.datetime):
    return "fecha_hora"
  if isinstance(value, dt.date):
    return "fecha"
  return "texto"


def clean_cell_value(value):
  if isinstance(value, dt.datetime):
    return value.isoformat(sep=" ")
  if isinstance(value, dt.date):
    return value.isoformat()
  if isinstance(value, str):
    return re.sub(r"\s+", " ", value).strip()
  return value


def extract_xlsx(path: Path, resource: dict, writer: csv.DictWriter) -> list[dict]:
  wb = load_workbook(path, read_only=True, data_only=True)
  sheets = []
  rid = resource_id(resource)
  for ws in wb.worksheets:
    non_empty = 0
    for row_idx, values in enumerate(ws.iter_rows(values_only=True), start=1):
      for col_idx, value in enumerate(values, start=1):
        if value is None or value == "":
          continue
        non_empty += 1
        col_letter = get_column_letter(col_idx)
        writer.writerow({
          "ID recurso": rid,
          "Título": resource["Título"],
          "Categoría": resource["Categoría"],
          "Formato": resource["Formato"],
          "URL fuente": resource["URL fuente"],
          "Hoja": ws.title,
          "Fila": row_idx,
          "Columna": col_letter,
          "Celda": f"{col_letter}{row_idx}",
          "Valor": clean_cell_value(value),
          "Tipo valor": value_type(value),
        })
    sheets.append({
      "ID recurso": rid,
      "Título": resource["Título"],
      "Categoría": resource["Categoría"],
      "Formato": resource["Formato"],
      "URL fuente": resource["URL fuente"],
      "Hoja": ws.title,
      "Filas detectadas": ws.max_row,
      "Columnas detectadas": ws.max_column,
      "Celdas con valor": non_empty,
      "Estado extracción": "OK",
      "Error": "",
    })
  wb.close()
  return sheets


def extract_xls(path: Path, resource: dict, writer: csv.DictWriter) -> list[dict]:
  if xlrd is None:
    raise RuntimeError("xlrd is required to read .xls files")
  book = xlrd.open_workbook(path, on_demand=True)
  sheets = []
  rid = resource_id(resource)
  for sheet_name in book.sheet_names():
    sheet = book.sheet_by_name(sheet_name)
    non_empty = 0
    for row_idx in range(sheet.nrows):
      for col_idx in range(sheet.ncols):
        value = sheet.cell_value(row_idx, col_idx)
        if value in ("", None):
          continue
        non_empty += 1
        col_letter = get_column_letter(col_idx + 1)
        writer.writerow({
          "ID recurso": rid,
          "Título": resource["Título"],
          "Categoría": resource["Categoría"],
          "Formato": resource["Formato"],
          "URL fuente": resource["URL fuente"],
          "Hoja": sheet_name,
          "Fila": row_idx + 1,
          "Columna": col_letter,
          "Celda": f"{col_letter}{row_idx + 1}",
          "Valor": clean_cell_value(value),
          "Tipo valor": value_type(value),
        })
    sheets.append({
      "ID recurso": rid,
      "Título": resource["Título"],
      "Categoría": resource["Categoría"],
      "Formato": resource["Formato"],
      "URL fuente": resource["URL fuente"],
      "Hoja": sheet_name,
      "Filas detectadas": sheet.nrows,
      "Columnas detectadas": sheet.ncols,
      "Celdas con valor": non_empty,
      "Estado extracción": "OK",
      "Error": "",
    })
  book.release_resources()
  return sheets


def extract_tabular_values(resources: list[dict]) -> tuple[list[dict], int]:
  tabular = [row for row in resources if row["Tipo de recurso"] == "Dato tabular"]
  output_path = CSV_DIR / "ove_ine_venezuela_celdas_tabulares.csv.gz"
  sheet_index = []
  total_cells = 0
  with tempfile.TemporaryDirectory() as tmpdir:
    tmpdir_path = Path(tmpdir)
    with gzip.open(output_path, "wt", newline="", encoding="utf-8") as handle:
      writer = csv.DictWriter(handle, fieldnames=CELL_FIELDS, delimiter=";", lineterminator="\n")
      writer.writeheader()
      for index, resource in enumerate(tabular, start=1):
        if index == 1 or index % 10 == 0 or index == len(tabular):
          print(f"INE tabular extraction {index}/{len(tabular)}: {resource['Título']}", file=sys.stderr, flush=True)
        rid = resource_id(resource)
        fmt = resource["Formato"].lower()
        local_path = tmpdir_path / f"{rid}.{fmt}"
        try:
          local_path.write_bytes(fetch_bytes(resource["URL fuente"]))
          if fmt == "xlsx":
            sheets = extract_xlsx(local_path, resource, writer)
          elif fmt == "xls":
            sheets = extract_xls(local_path, resource, writer)
          else:
            sheets = []
          total_cells += sum(sheet["Celdas con valor"] for sheet in sheets)
          sheet_index.extend(sheets)
        except Exception as exc:  # noqa: BLE001 - catalog extraction should continue.
          sheet_index.append({
            "ID recurso": rid,
            "Título": resource["Título"],
            "Categoría": resource["Categoría"],
            "Formato": resource["Formato"],
            "URL fuente": resource["URL fuente"],
            "Hoja": "",
            "Filas detectadas": 0,
            "Columnas detectadas": 0,
            "Celdas con valor": 0,
            "Estado extracción": "Error",
            "Error": str(exc),
          })
  return sheet_index, total_cells



def load_resources() -> list[dict]:
  RAW_DIR.mkdir(parents=True, exist_ok=True)
  html_text = fetch_text(HOME_URL)
  (RAW_DIR / "ine_venezuela_home.html").write_text(html_text, encoding="utf-8")
  records = {}
  for match in re.finditer(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>(.*?)</a>', html_text, flags=re.I | re.S):
    url = urllib.parse.urljoin(HOME_URL, html.unescape(match.group(1)))
    fmt = extension(url)
    if fmt not in {"xls", "xlsx", "csv", "pdf"}:
      continue
    if urllib.parse.urlparse(url).netloc != "ine.gob.ve":
      continue
    title = clean_text(match.group(2)) or urllib.parse.unquote(Path(urllib.parse.urlparse(url).path).name)
    if url in records:
      if len(title) > len(records[url]["Título"]):
        records[url]["Título"] = title
      continue
    year, month = publication_date_parts(url)
    records[url] = {
      "Fuente": SOURCE_NAME,
      "País": "Venezuela",
      "Título": title,
      "Categoría": infer_category(title, url),
      "Tipo de recurso": resource_type(fmt),
      "Formato": fmt.upper(),
      "Año publicación": year,
      "Mes publicación": month,
      "Archivo original": urllib.parse.unquote(Path(urllib.parse.urlparse(url).path).name),
      "URL fuente": url,
      "URL portal": HOME_URL,
    }
  return sorted(records.values(), key=lambda row: (row["Categoría"], row["Formato"], row["Título"], row["URL fuente"]))


def write_outputs(resources: list[dict], generated_at: str) -> None:
  for directory in (CATALOG_DIR, CSV_DIR, JSON_DIR, EXCEL_DIR):
    directory.mkdir(parents=True, exist_ok=True)
  sheet_index, total_cells = extract_tabular_values(resources)

  csv_path = CATALOG_DIR / "catalogo_dataset_web_ove_ine_venezuela.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=CATALOG_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(resources)

  # Same catalog under csv/ for consistency with source pages that expose data downloads.
  csv_copy = CSV_DIR / "ove_ine_venezuela_catalogo_recursos.csv"
  csv_copy.write_text(csv_path.read_text(encoding="utf-8"), encoding="utf-8")

  tabular = [row for row in resources if row["Tipo de recurso"] == "Dato tabular"]
  documents = [row for row in resources if row["Tipo de recurso"] == "Documento"]
  payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "INE Venezuela - catálogo de recursos oficiales",
      "Fecha generación": generated_at,
      "Fuente": SOURCE_NAME,
      "URL portal": HOME_URL,
      "País": "Venezuela",
      "Número de recursos": len(resources),
      "Recursos tabulares": len(tabular),
      "Documentos": len(documents),
      "Hojas extraídas": sum(1 for row in sheet_index if row["Estado extracción"] == "OK"),
      "Celdas con valor extraídas": total_cells,
      "Archivo valores CSV.GZ": "assets/data/ine/csv/ove_ine_venezuela_celdas_tabulares.csv.gz",
    },
    "catalogo": resources,
    "indice_hojas": sheet_index,
  }
  (CATALOG_DIR / "ine-catalog.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
  (JSON_DIR / "ove_ine_venezuela_catalogo_recursos.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb = Workbook()
  ws = wb.active
  ws.title = "catalogo"
  format_sheet(ws, "INE Venezuela", "Catálogo de recursos oficiales")
  write_table(ws, CATALOG_FIELDS, resources)
  meta = wb.create_sheet("metadatos")
  format_sheet(meta, "Metadatos", "INE Venezuela")
  write_key_values(meta, payload["metadatos"].items())
  wb.save(CATALOG_DIR / "catalogo_dataset_web_ove_ine_venezuela.xlsx")

  wb_public = Workbook()
  ws_public = wb_public.active
  ws_public.title = "catalogo"
  format_sheet(ws_public, "INE Venezuela", "Catálogo de recursos oficiales")
  write_table(ws_public, CATALOG_FIELDS, resources)
  meta_public = wb_public.create_sheet("metadatos")
  format_sheet(meta_public, "Metadatos", "INE Venezuela")
  write_key_values(meta_public, payload["metadatos"].items())
  wb_public.save(EXCEL_DIR / "ove_ine_venezuela_catalogo_recursos.xlsx")

  sheets_csv = CSV_DIR / "ove_ine_venezuela_indice_hojas.csv"
  with sheets_csv.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=SHEET_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(sheet_index)

  wb_sheets = Workbook()
  ws_sheets = wb_sheets.active
  ws_sheets.title = "indice_hojas"
  format_sheet(ws_sheets, "INE Venezuela", "Índice de hojas extraídas")
  write_table(ws_sheets, SHEET_FIELDS, sheet_index)
  meta_sheets = wb_sheets.create_sheet("metadatos")
  format_sheet(meta_sheets, "Metadatos", "Valores tabulares INE")
  write_key_values(meta_sheets, payload["metadatos"].items())
  wb_sheets.save(EXCEL_DIR / "ove_ine_venezuela_indice_hojas.xlsx")


def main() -> int:
  generated_at = today()
  resources = load_resources()
  write_outputs(resources, generated_at)
  print(json.dumps({
    "dataset": "INE Venezuela - catálogo de recursos oficiales",
    "generated_at": generated_at,
    "resources": len(resources),
    "tabular": sum(1 for row in resources if row["Tipo de recurso"] == "Dato tabular"),
    "documents": sum(1 for row in resources if row["Tipo de recurso"] == "Documento"),
    "sheets": sum(1 for row in json.loads((JSON_DIR / "ove_ine_venezuela_catalogo_recursos.json").read_text(encoding="utf-8"))["indice_hojas"] if row["Estado extracción"] == "OK"),
    "cells": json.loads((JSON_DIR / "ove_ine_venezuela_catalogo_recursos.json").read_text(encoding="utf-8"))["metadatos"]["Celdas con valor extraídas"],
    "output": str(INE_ROOT),
  }, ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
