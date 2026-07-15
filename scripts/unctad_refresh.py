#!/usr/bin/env python3
"""Refresh UNCTADstat Venezuela datasets from public bulk downloads."""

from __future__ import annotations

import csv
import datetime as dt
import gzip
import io
import json
import os
import re
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

import py7zr
from openpyxl import Workbook

from ove_excel_format import format_sheet, write_key_values, write_table


ROOT = Path(__file__).resolve().parents[1]
UNCTAD_ROOT = ROOT / "assets" / "data" / "unctad"
RAW_DIR = UNCTAD_ROOT / "raw"
CSV_DIR = UNCTAD_ROOT / "csv"
JSON_DIR = UNCTAD_ROOT / "json"
EXCEL_DIR = UNCTAD_ROOT / "excel"
CATALOG_DIR = UNCTAD_ROOT / "catalog"

DATACENTER_URL = "https://unctadstat.unctad.org/datacentre/"
API_ROOT = "https://unctadstat-api.unctad.org"
DATACENTER_API = f"{API_ROOT}/api/datacenter/en"
COUNTRY_NAME = "Venezuela (Bolivarian Rep. of)"
COUNTRY_CODE = "862"
COUNTRY_ISO3 = "VEN"
USER_AGENT = "OVE UNCTADstat refresh/1.0 (+https://ove-web-github.vercel.app/)"
MAX_BULK_BYTES = int(float(os.environ.get("UNCTAD_MAX_BULK_MB", "20")) * 1024 * 1024)
EXCEL_MAX_ROWS = 1_048_000

VALUE_FIELDS = [
  "País",
  "Código país",
  "Fuente",
  "Reporte UNCTAD",
  "ID reporte",
  "Título reporte",
  "Categoría",
  "Ruta temática",
  "Archivo bulk",
  "Archivo CSV interno",
  "Periodo",
  "Año",
  "Medida",
  "Valor",
  "Dimensiones",
  "Fila original JSON",
  "Última actualización reporte",
  "URL bulk",
  "URL fuente",
]

CATALOG_FIELDS = [
  "Fuente",
  "País",
  "Reporte UNCTAD",
  "ID reporte",
  "Título reporte",
  "Categoría",
  "Ruta temática",
  "Palabras clave",
  "Última actualización reporte",
  "Número de archivos bulk",
  "Archivos procesados",
  "Archivos omitidos por tamaño",
  "Tamaño bulk total bytes",
  "Número de registros",
  "Primer periodo",
  "Último periodo",
  "Estado descarga",
  "Error descarga",
  "Archivo valores CSV.GZ",
  "Archivo catálogo JSON",
  "Archivo catálogo Excel",
  "URL fuente",
]

BULK_FIELDS = [
  "Fuente",
  "Reporte UNCTAD",
  "Título reporte",
  "Archivo bulk",
  "Nombre archivo",
  "Tamaño bytes",
  "Estado",
  "Registros extraídos",
  "Valores extraídos",
  "URL bulk",
]


def today() -> str:
  return dt.date.today().isoformat()


def request_bytes(url: str, timeout: int = 120, attempts: int = 3) -> bytes:
  last_error: Exception | None = None
  for attempt in range(1, attempts + 1):
    try:
      request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
      with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()
    except (urllib.error.URLError, urllib.error.HTTPError, TimeoutError) as exc:
      last_error = exc
      if attempt < attempts:
        time.sleep(0.8 * attempt)
  raise RuntimeError(str(last_error))


def request_json(url: str, timeout: int = 120, attempts: int = 3) -> Any:
  return json.loads(request_bytes(url, timeout=timeout, attempts=attempts).decode("utf-8-sig"))


def safe_float(value: object) -> float | None:
  if value in (None, ""):
    return None
  text = str(value).strip().replace("\u00a0", "")
  if not text or text in {".", "..", "...", "-"}:
    return None
  try:
    return float(text.replace(",", ""))
  except ValueError:
    return None


def infer_year(value: object) -> int | None:
  match = re.search(r"\b(18|19|20|21)\d{2}\b", str(value or ""))
  return int(match.group(0)) if match else None


def period_sort_key(value: object) -> tuple[int, str]:
  year = infer_year(value)
  return (year or -9999, str(value or ""))


def clean_text(value: object) -> str:
  return re.sub(r"\s+", " ", str(value or "")).strip()


def bulkfiles_url(report_name: str) -> str:
  return f"{API_ROOT}/api/reportMetadata/{urllib.parse.quote(report_name)}/bulkfiles/en"


def metadata_url(report_name: str) -> str:
  return f"{API_ROOT}/api/reportMetadata/{urllib.parse.quote(report_name)}/en"


def bulk_download_url(report_name: str, blob_name: str) -> str:
  return f"{API_ROOT}/bulkdownload/{urllib.parse.quote(report_name)}/{urllib.parse.quote(blob_name)}"


def load_reports() -> list[dict]:
  tree = request_json(DATACENTER_API, timeout=120)
  RAW_DIR.mkdir(parents=True, exist_ok=True)
  (RAW_DIR / "unctadstat_datacenter_tree_en.json").write_text(
    json.dumps(tree, ensure_ascii=False, indent=2) + "\n",
    encoding="utf-8",
  )
  reports: list[dict] = []

  def walk(nodes: list[dict], path: list[str]) -> None:
    for node in nodes:
      current_path = path + ([node.get("title", "")] if node.get("title") else [])
      for report in node.get("reports") or []:
        reports.append({**report, "Ruta temática": " > ".join(current_path)})
      walk(node.get("subFolders") or [], current_path)

  walk(tree, [])
  return reports


def load_report_metadata(report_name: str) -> dict:
  try:
    return request_json(metadata_url(report_name), timeout=90)
  except Exception:
    return {}


def load_bulkfiles(report_name: str) -> list[dict]:
  try:
    payload = request_json(bulkfiles_url(report_name), timeout=90)
    return payload if isinstance(payload, list) else []
  except Exception:
    return []


def metadata_fields(metadata: dict) -> tuple[set[str], str]:
  defaults = metadata.get("defaults") or {}
  axes = (defaults.get("rowAxe") or []) + (defaults.get("colAxe") or []) + (defaults.get("pageAxe") or [])
  dimension_fields = {item.get("field") for item in axes if item.get("field")}
  time_field = next((item.get("field") for item in axes if item.get("isTime") and item.get("field")), "")
  return set(dimension_fields), time_field


def is_venezuela_row(row: dict[str, str]) -> bool:
  for key, value in row.items():
    text = clean_text(value).lower()
    if "venezuela" in text:
      return True
    if key.lower() in {
      "economy", "reporter", "partner", "economy code", "reporter code", "partner code",
      "country", "country code", "geo", "geo code",
    } and text in {COUNTRY_CODE, "0862", "ven"}:
      return True
  return False


def is_dimension_column(field: str, dimension_fields: set[str]) -> bool:
  lower = field.lower()
  if field in dimension_fields or field.replace(" Label", "") in dimension_fields:
    return True
  if lower.endswith(" label") or lower.endswith(" code") or lower in {"year", "period", "quarter", "month"}:
    return True
  if lower in {"economy", "reporter", "partner", "flow", "direction", "mode", "product", "service", "category"}:
    return True
  if "footnote" in lower or "missing value" in lower:
    return True
  return False


def dimensions_json(row: dict[str, str], measure: str, dimension_fields: set[str]) -> str:
  dimensions = {
    key: value for key, value in row.items()
    if value not in ("", None)
    and key != measure
    and is_dimension_column(key, dimension_fields)
  }
  return json.dumps(dimensions, ensure_ascii=False, sort_keys=True)


def original_row_json(row: dict[str, str]) -> str:
  return json.dumps({key: value for key, value in row.items() if value not in ("", None)}, ensure_ascii=False, sort_keys=True)


def detect_delimiter(sample: str) -> str:
  try:
    return csv.Sniffer().sniff(sample).delimiter
  except csv.Error:
    return ","


def extract_csv_files(archive_bytes: bytes) -> list[tuple[str, str]]:
  files: list[tuple[str, str]] = []
  with tempfile.TemporaryDirectory() as temp_dir:
    archive_path = Path(temp_dir) / "download.7z"
    out_dir = Path(temp_dir) / "out"
    out_dir.mkdir()
    archive_path.write_bytes(archive_bytes)
    with py7zr.SevenZipFile(archive_path, "r") as archive:
      archive.extractall(out_dir)
    for path in out_dir.rglob("*.csv"):
      text = path.read_text(encoding="utf-8-sig", errors="replace")
      files.append((path.name, text))
  return files


def parse_bulk_file(report: dict, metadata: dict, bulk_file: dict) -> tuple[list[dict], int, str]:
  report_name = report["reportName"]
  blob_name = bulk_file["fileBlobName"]
  url = bulk_download_url(report_name, blob_name)
  rows: list[dict] = []
  original_rows = 0
  try:
    archive_bytes = request_bytes(url, timeout=240)
    csv_files = extract_csv_files(archive_bytes)
    dimension_fields, time_field = metadata_fields(metadata)
    for internal_name, text in csv_files:
      sample = "\n".join(text.splitlines()[:5])
      delimiter = detect_delimiter(sample)
      reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
      for raw_row in reader:
        if not is_venezuela_row(raw_row):
          continue
        original_rows += 1
        period = raw_row.get(time_field) or raw_row.get("Year") or raw_row.get("Period") or raw_row.get("Quarter") or raw_row.get("Month") or ""
        original = original_row_json(raw_row)
        for field, raw_value in raw_row.items():
          if is_dimension_column(field, dimension_fields):
            continue
          value = safe_float(raw_value)
          if value is None:
            continue
          rows.append({
            "País": COUNTRY_NAME,
            "Código país": COUNTRY_CODE,
            "Fuente": "UNCTADstat - UNCTAD",
            "Reporte UNCTAD": report_name,
            "ID reporte": report.get("id"),
            "Título reporte": report.get("reportTitle", ""),
            "Categoría": report.get("category", ""),
            "Ruta temática": report.get("Ruta temática", ""),
            "Archivo bulk": blob_name,
            "Archivo CSV interno": internal_name,
            "Periodo": period,
            "Año": infer_year(period),
            "Medida": field,
            "Valor": value,
            "Dimensiones": dimensions_json(raw_row, field, dimension_fields),
            "Fila original JSON": original,
            "Última actualización reporte": report.get("lastUpdatedDate") or metadata.get("lastUpdated", ""),
            "URL bulk": url,
            "URL fuente": DATACENTER_URL,
          })
    return rows, original_rows, ""
  except Exception as exc:  # noqa: BLE001 - keep source failures in catalog.
    return [], original_rows, str(exc)[:500]


def process_reports(reports: list[dict]) -> tuple[list[dict], list[dict], list[dict]]:
  catalog: list[dict] = []
  bulk_catalog: list[dict] = []
  values: list[dict] = []
  for index, report in enumerate(reports, start=1):
    report_name = report["reportName"]
    metadata = load_report_metadata(report_name)
    bulk_files = load_bulkfiles(report_name)
    processed_files = 0
    skipped_files = 0
    report_values = 0
    report_error = ""
    periods: list[str] = []
    for bulk_file in bulk_files:
      size = int(bulk_file.get("fileSize") or 0)
      blob = bulk_file.get("fileBlobName", "")
      url = bulk_download_url(report_name, blob)
      if size > MAX_BULK_BYTES:
        skipped_files += 1
        bulk_catalog.append({
          "Fuente": "UNCTADstat - UNCTAD",
          "Reporte UNCTAD": report_name,
          "Título reporte": report.get("reportTitle", ""),
          "Archivo bulk": blob,
          "Nombre archivo": bulk_file.get("fileName", ""),
          "Tamaño bytes": size,
          "Estado": f"Omitido por tamaño > {MAX_BULK_BYTES} bytes",
          "Registros extraídos": 0,
          "Valores extraídos": 0,
          "URL bulk": url,
        })
        continue
      file_rows, original_rows, error = parse_bulk_file(report, metadata, bulk_file)
      if error:
        report_error = "; ".join([item for item in (report_error, f"{blob}: {error}") if item])
      processed_files += 1
      report_values += len(file_rows)
      values.extend(file_rows)
      periods.extend([row["Periodo"] for row in file_rows if row.get("Periodo")])
      bulk_catalog.append({
        "Fuente": "UNCTADstat - UNCTAD",
        "Reporte UNCTAD": report_name,
        "Título reporte": report.get("reportTitle", ""),
        "Archivo bulk": blob,
        "Nombre archivo": bulk_file.get("fileName", ""),
        "Tamaño bytes": size,
        "Estado": "Procesado" if not error else "Error",
        "Registros extraídos": original_rows,
        "Valores extraídos": len(file_rows),
        "URL bulk": url,
      })
    status = "Con valores para Venezuela" if report_values else "Sin valores descargados para Venezuela"
    if skipped_files and not processed_files:
      status = "Solo catalogado por tamaño de bulk"
    if report_error and not report_values:
      status = "Error de descarga"
    catalog.append({
      "Fuente": "UNCTADstat - UNCTAD",
      "País": COUNTRY_NAME,
      "Reporte UNCTAD": report_name,
      "ID reporte": report.get("id"),
      "Título reporte": report.get("reportTitle", ""),
      "Categoría": report.get("category", ""),
      "Ruta temática": report.get("Ruta temática", ""),
      "Palabras clave": report.get("keywords", ""),
      "Última actualización reporte": report.get("lastUpdatedDate") or metadata.get("lastUpdated", ""),
      "Número de archivos bulk": len(bulk_files),
      "Archivos procesados": processed_files,
      "Archivos omitidos por tamaño": skipped_files,
      "Tamaño bulk total bytes": sum(int(item.get("fileSize") or 0) for item in bulk_files),
      "Número de registros": report_values,
      "Primer periodo": min(periods, key=period_sort_key) if periods else "",
      "Último periodo": max(periods, key=period_sort_key) if periods else "",
      "Estado descarga": status,
      "Error descarga": report_error[:500],
      "Archivo valores CSV.GZ": "assets/data/unctad/csv/ove_unctadstat_venezuela_valores.csv.gz",
      "Archivo catálogo JSON": "assets/data/unctad/catalog/unctad-catalog.json",
      "Archivo catálogo Excel": "assets/data/unctad/catalog/catalogo_dataset_web_ove_unctadstat.xlsx",
      "URL fuente": DATACENTER_URL,
    })
    if index % 10 == 0:
      print(json.dumps({
        "dataset": "UNCTADstat - Venezuela",
        "processed_reports": index,
        "reports": len(reports),
        "values_partial": len(values),
      }, ensure_ascii=False), flush=True)
  return catalog, bulk_catalog, values


def write_outputs(catalog: list[dict], bulk_catalog: list[dict], values: list[dict], generated_at: str) -> dict:
  for directory in (CSV_DIR, JSON_DIR, EXCEL_DIR, CATALOG_DIR):
    directory.mkdir(parents=True, exist_ok=True)

  catalog.sort(key=lambda row: (row["Número de registros"] == 0, row["Ruta temática"], row["Título reporte"]))
  values.sort(key=lambda row: (row["Reporte UNCTAD"], period_sort_key(row["Periodo"]), row["Medida"]))

  values_path = CSV_DIR / "ove_unctadstat_venezuela_valores.csv.gz"
  with gzip.open(values_path, "wt", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=VALUE_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(values)

  catalog_csv = CATALOG_DIR / "catalogo_dataset_web_ove_unctadstat.csv"
  with catalog_csv.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=CATALOG_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(catalog)

  bulk_csv = CSV_DIR / "ove_unctadstat_bulkfiles.csv"
  with bulk_csv.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=BULK_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(bulk_catalog)

  years = [row["Año"] for row in values if row.get("Año")]
  metadata = {
    "Organización": "Observatorio Venezolano de Economía",
    "Dataset": "UNCTADstat - Venezuela",
    "Fecha generación": generated_at,
    "Fuente": "UNCTADstat - UNCTAD",
    "URL fuente": DATACENTER_URL,
    "API catálogo": DATACENTER_API,
    "País": COUNTRY_NAME,
    "Código país UNCTAD": COUNTRY_CODE,
    "Código ISO3": COUNTRY_ISO3,
    "Número de reportes catalogados": len(catalog),
    "Número de reportes con valores": sum(1 for row in catalog if row["Número de registros"] > 0),
    "Número de archivos bulk catalogados": len(bulk_catalog),
    "Número de valores extraídos": len(values),
    "Primer año": min(years) if years else "",
    "Último año": max(years) if years else "",
    "Límite automático de descarga por archivo MB": round(MAX_BULK_BYTES / 1024 / 1024, 2),
    "Nota": "Los bulk files oficiales que superan el límite operativo quedan catalogados con URL oficial y no se sustituyen por archivos vacíos.",
  }
  payload = {"metadatos": metadata, "catalogo": catalog, "archivos_bulk": bulk_catalog}
  (CATALOG_DIR / "unctad-catalog.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
  (JSON_DIR / "ove_unctadstat_venezuela_metadata.json").write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb_catalog = Workbook()
  ws = wb_catalog.active
  ws.title = "catalogo"
  format_sheet(ws, "UNCTADstat - Venezuela", "Catálogo de reportes y estado de descarga")
  write_table(ws, CATALOG_FIELDS, catalog)
  bulk_ws = wb_catalog.create_sheet("bulkfiles")
  format_sheet(bulk_ws, "UNCTADstat - bulk files", "Archivos oficiales catalogados")
  write_table(bulk_ws, BULK_FIELDS, bulk_catalog)
  meta_ws = wb_catalog.create_sheet("metadatos")
  format_sheet(meta_ws, "Metadatos", "UNCTADstat - Venezuela")
  write_key_values(meta_ws, metadata.items())
  wb_catalog.save(CATALOG_DIR / "catalogo_dataset_web_ove_unctadstat.xlsx")

  if values and len(values) < EXCEL_MAX_ROWS:
    wb_values = Workbook()
    value_ws = wb_values.active
    value_ws.title = "valores"
    format_sheet(value_ws, "UNCTADstat - Venezuela", "Valores extraídos")
    write_table(value_ws, VALUE_FIELDS, values)
    meta_values = wb_values.create_sheet("metadatos")
    format_sheet(meta_values, "Metadatos", "UNCTADstat - Venezuela")
    write_key_values(meta_values, metadata.items())
    wb_values.save(EXCEL_DIR / "ove_unctadstat_venezuela_valores.xlsx")

  return metadata


def main() -> int:
  generated_at = today()
  reports = load_reports()
  catalog, bulk_catalog, values = process_reports(reports)
  metadata = write_outputs(catalog, bulk_catalog, values, generated_at)
  print(json.dumps({
    "dataset": "UNCTADstat - Venezuela",
    "generated_at": generated_at,
    "reports_cataloged": metadata["Número de reportes catalogados"],
    "reports_with_values": metadata["Número de reportes con valores"],
    "bulk_files": metadata["Número de archivos bulk catalogados"],
    "values": metadata["Número de valores extraídos"],
    "output": str(UNCTAD_ROOT),
  }, ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
