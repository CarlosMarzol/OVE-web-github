#!/usr/bin/env python3
"""Refresh IMF WEO Venezuela datasets from the public SDMX 3.0 API."""

from __future__ import annotations

import csv
import datetime as dt
import io
import json
import urllib.parse
import urllib.request
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook

from ove_excel_format import format_sheet, write_key_values, write_table


ROOT = Path(__file__).resolve().parents[1]
IMF_ROOT = ROOT / "assets" / "data" / "imf"
RAW_DIR = IMF_ROOT / "raw"
CSV_DIR = IMF_ROOT / "csv"
JSON_DIR = IMF_ROOT / "json"
EXCEL_DIR = IMF_ROOT / "excel"
CATALOG_DIR = IMF_ROOT / "catalog"

STRUCTURE_URL = "https://api.imf.org/external/sdmx/3.0/structure/dataflow/IMF.RES/WEO/9.0.0?references=all"
DATA_URL = "https://api.imf.org/external/sdmx/3.0/data/dataflow/IMF.RES/WEO/~/*"
SOURCE_URL = "https://data.imf.org/en/datasets/IMF.RES:WEO"

DATA_FIELDS = [
  "País",
  "Código país",
  "Indicador",
  "Código indicador",
  "Descripción indicador",
  "Frecuencia",
  "Periodo",
  "Año",
  "Valor",
  "Unidad",
  "Escala",
  "Decimales",
  "Tipo dato",
  "Fecha actualización país",
  "Fuente",
  "URL fuente",
]

CATALOG_FIELDS = [
  "Fuente",
  "País",
  "Código indicador",
  "Indicador",
  "Descripción indicador",
  "Unidad",
  "Escala",
  "Frecuencia",
  "Número de registros",
  "Primer periodo",
  "Último periodo",
  "Último valor disponible",
  "Fecha actualización país",
  "Archivo CSV",
  "Archivo JSON",
  "Archivo Excel",
]


def today() -> str:
  return dt.date.today().isoformat()


def fetch_bytes(url: str, accept: str) -> bytes:
  request = urllib.request.Request(url, headers={"Accept": accept, "User-Agent": "OVE IMF refresh/1.0"})
  with urllib.request.urlopen(request, timeout=90) as response:
    return response.read()


def code_label_maps(structure: dict) -> dict[str, dict[str, dict]]:
  maps = {}
  for codelist in structure["data"].get("codelists", []):
    mapping = {}
    for code in codelist.get("codes", []):
      mapping[code["id"]] = {
        "label": code.get("name") or code.get("names", {}).get("en") or code["id"],
        "description": code.get("description") or code.get("descriptions", {}).get("en") or "",
      }
    maps[codelist["id"]] = mapping
  return maps


def safe_float(value: str | None) -> float | None:
  if value in (None, ""):
    return None
  try:
    return float(value)
  except ValueError:
    return None


def period_key(value: str) -> int:
  try:
    return int(str(value)[:4])
  except ValueError:
    return 0


def load_venezuela_rows() -> tuple[list[dict], dict[str, dict[str, dict]], str]:
  params = urllib.parse.urlencode({"c[COUNTRY]": "VEN"})
  data_url = f"{DATA_URL}?{params}"
  structure = json.loads(fetch_bytes(STRUCTURE_URL, "application/vnd.sdmx.structure+json;version=2.0.0").decode("utf-8"))
  raw_csv = fetch_bytes(data_url, "text/csv")
  RAW_DIR.mkdir(parents=True, exist_ok=True)
  RAW_DIR.joinpath("ove_fmi_weo_venezuela_raw.csv").write_bytes(raw_csv)
  text = raw_csv.decode("utf-8-sig")
  rows = list(csv.DictReader(io.StringIO(text)))
  return rows, code_label_maps(structure), data_url


def normalize_rows(raw_rows: list[dict], maps: dict[str, dict[str, dict]]) -> list[dict]:
  indicators = maps.get("CL_WEO_INDICATOR", {})
  countries = maps.get("CL_WEO_COUNTRY", {}) or maps.get("CL_COUNTRY", {})
  units = maps.get("CL_UNIT", {})
  freqs = maps.get("CL_FREQ", {})
  overlaps = maps.get("CL_OVERLAP", {})
  output = []
  for row in raw_rows:
    if row.get("COUNTRY") != "VEN" or not row.get("TIME_PERIOD"):
      continue
    indicator_code = row.get("INDICATOR", "")
    country_code = row.get("COUNTRY", "")
    unit_code = row.get("UNIT", "")
    freq_code = row.get("FREQUENCY", "")
    overlap_code = row.get("OVERLAP", "")
    indicator = indicators.get(indicator_code, {})
    country = countries.get(country_code, {})
    unit = units.get(unit_code, {})
    freq = freqs.get(freq_code, {})
    overlap = overlaps.get(overlap_code, {})
    period = row.get("TIME_PERIOD", "")
    output.append({
      "País": country.get("label") or "Venezuela",
      "Código país": country_code,
      "Indicador": indicator.get("label") or indicator_code,
      "Código indicador": indicator_code,
      "Descripción indicador": indicator.get("description", ""),
      "Frecuencia": freq.get("label") or freq_code,
      "Periodo": period,
      "Año": period_key(period),
      "Valor": safe_float(row.get("OBS_VALUE")),
      "Unidad": unit.get("label") or unit_code,
      "Escala": row.get("SCALE", ""),
      "Decimales": row.get("DECIMALS_DISPLAYED", ""),
      "Tipo dato": overlap.get("label") or overlap_code,
      "Fecha actualización país": row.get("COUNTRY_UPDATE_DATE", ""),
      "Fuente": "FMI - World Economic Outlook",
      "URL fuente": SOURCE_URL,
    })
  output.sort(key=lambda item: (item["Código indicador"], item["Año"]))
  return output


def write_data_outputs(rows: list[dict], generated_at: str, data_url: str) -> None:
  for directory in (CSV_DIR, JSON_DIR, EXCEL_DIR, CATALOG_DIR):
    directory.mkdir(parents=True, exist_ok=True)
  csv_path = CSV_DIR / "ove_fmi_weo_venezuela.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=DATA_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)

  payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "FMI - World Economic Outlook - Venezuela",
      "Fecha generación": generated_at,
      "Fuente": "FMI - World Economic Outlook",
      "URL fuente": SOURCE_URL,
      "API": data_url,
      "País": "Venezuela",
      "Número de indicadores": len({row["Código indicador"] for row in rows}),
      "Número de registros": len(rows),
    },
    "datos": rows,
  }
  (JSON_DIR / "ove_fmi_weo_venezuela.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb = Workbook()
  ws = wb.active
  ws.title = "datos"
  format_sheet(ws, "FMI - World Economic Outlook", "Venezuela")
  write_table(ws, DATA_FIELDS, rows)
  meta = wb.create_sheet("metadatos")
  format_sheet(meta, "Metadatos", "FMI - World Economic Outlook")
  write_key_values(meta, payload["metadatos"].items())
  wb.save(EXCEL_DIR / "ove_fmi_weo_venezuela.xlsx")


def catalog_rows(rows: list[dict]) -> list[dict]:
  grouped: dict[str, list[dict]] = defaultdict(list)
  for row in rows:
    grouped[row["Código indicador"]].append(row)
  catalog = []
  for code, items in sorted(grouped.items()):
    values = [item for item in items if item["Valor"] is not None]
    latest = max(values, key=lambda item: item["Año"]) if values else items[-1]
    catalog.append({
      "Fuente": "FMI - World Economic Outlook",
      "País": "Venezuela",
      "Código indicador": code,
      "Indicador": items[0]["Indicador"],
      "Descripción indicador": items[0]["Descripción indicador"],
      "Unidad": items[0]["Unidad"],
      "Escala": items[0]["Escala"],
      "Frecuencia": items[0]["Frecuencia"],
      "Número de registros": len(items),
      "Primer periodo": min((item["Periodo"] for item in items), key=period_key),
      "Último periodo": max((item["Periodo"] for item in items), key=period_key),
      "Último valor disponible": latest.get("Valor"),
      "Fecha actualización país": latest.get("Fecha actualización país"),
      "Archivo CSV": "assets/data/imf/csv/ove_fmi_weo_venezuela.csv",
      "Archivo JSON": "assets/data/imf/json/ove_fmi_weo_venezuela.json",
      "Archivo Excel": "assets/data/imf/excel/ove_fmi_weo_venezuela.xlsx",
    })
  return catalog


def write_catalog(catalog: list[dict], generated_at: str) -> None:
  csv_path = CATALOG_DIR / "catalogo_dataset_web_ove_fmi_weo.csv"
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=CATALOG_FIELDS, delimiter=";", lineterminator="\n")
    writer.writeheader()
    writer.writerows(catalog)
  payload = {
    "metadatos": {
      "Organización": "Observatorio Venezolano de Economía",
      "Dataset": "FMI - World Economic Outlook - Venezuela",
      "Fecha generación": generated_at,
      "Fuente": "FMI - World Economic Outlook",
      "País": "Venezuela",
      "Número de indicadores": len(catalog),
      "Número de registros": sum(row["Número de registros"] for row in catalog),
    },
    "catalogo": catalog,
  }
  (CATALOG_DIR / "imf-catalog.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

  wb = Workbook()
  ws = wb.active
  ws.title = "catalogo"
  format_sheet(ws, "FMI - World Economic Outlook", "Catálogo Venezuela")
  write_table(ws, CATALOG_FIELDS, catalog)
  meta = wb.create_sheet("metadatos")
  format_sheet(meta, "Metadatos", "Catálogo FMI")
  write_key_values(meta, payload["metadatos"].items())
  wb.save(CATALOG_DIR / "catalogo_dataset_web_ove_fmi_weo.xlsx")


def main() -> int:
  generated_at = today()
  raw_rows, maps, data_url = load_venezuela_rows()
  rows = normalize_rows(raw_rows, maps)
  write_data_outputs(rows, generated_at, data_url)
  catalog = catalog_rows(rows)
  write_catalog(catalog, generated_at)
  print(json.dumps({
    "dataset": "FMI - World Economic Outlook - Venezuela",
    "generated_at": generated_at,
    "indicators": len(catalog),
    "records": len(rows),
    "output": str(IMF_ROOT),
  }, ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
