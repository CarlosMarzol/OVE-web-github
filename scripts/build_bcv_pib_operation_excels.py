#!/usr/bin/env python3
"""Build OVE-formatted BCV GDP workbooks for PIB operation pages."""

from __future__ import annotations

import csv
import datetime as dt
import json
import re
from pathlib import Path

try:
  import numpy as np
  for alias, value in (("float", float), ("bool", bool), ("int", int), ("object", object)):
    if not hasattr(np, alias):
      setattr(np, alias, value)
except Exception:
  pass

import xlrd
from openpyxl import Workbook, load_workbook

from ove_excel_format import format_sheet, write_key_values, write_table


ROOT = Path(__file__).resolve().parents[1]
BCV_ROOT = ROOT / "assets" / "data" / "bcv"
RAW_DIR = BCV_ROOT / "raw"
CSV_DIR = BCV_ROOT / "csv"
JSON_DIR = BCV_ROOT / "json"
EXCEL_DIR = BCV_ROOT / "excel"
CATALOG_DIR = BCV_ROOT / "catalog"
SOURCE = "Banco Central de Venezuela"
SOURCE_PAGE = "https://www.bcv.org.ve/estadisticas/producto-interno-bruto"
BASE_URL = "https://www.bcv.org.ve/sites/default/files/cuentas_macroeconomicas/"

FIELDS = [
  "dataset_id",
  "title",
  "table",
  "source_file",
  "base",
  "price_type",
  "measure",
  "frequency",
  "period",
  "year",
  "quarter",
  "classification",
  "component",
  "value",
  "unit",
  "status",
  "source",
  "source_url",
  "fetched_at",
]

QUARTERS = {
  "I": 1,
  "II": 2,
  "III": 3,
  "IV": 4,
}

DATASETS = {
  "bcv_pib_historico_anual": {
    "title": "Producto interno bruto histórico anual",
    "filename": "ove_bcv_pib_historico_anual.xlsx",
    "subtitle": "BCV - precios constantes y corrientes",
  },
  "bcv_pib_demanda_anual": {
    "title": "Producto interno bruto por componentes de demanda",
    "filename": "ove_bcv_pib_demanda_anual.xlsx",
    "subtitle": "BCV - enfoque de la demanda y producción",
  },
  "bcv_pib_sector_institucional_anual": {
    "title": "Producto interno bruto por sector institucional anual",
    "filename": "ove_bcv_pib_sector_institucional_anual.xlsx",
    "subtitle": "BCV - total, sector público, privado e impuestos",
  },
  "bcv_pib_sector_institucional_trimestral": {
    "title": "Producto interno bruto por sector institucional trimestral",
    "filename": "ove_bcv_pib_sector_institucional_trimestral.xlsx",
    "subtitle": "BCV - total, sector público, privado e impuestos",
  },
  "bcv_pib_actividad_economica_anual": {
    "title": "Producto interno bruto por actividad económica anual",
    "filename": "ove_bcv_pib_actividad_economica_anual.xlsx",
    "subtitle": "BCV - actividades económicas",
  },
  "bcv_pib_actividad_economica_trimestral": {
    "title": "Producto interno bruto por actividad económica trimestral",
    "filename": "ove_bcv_pib_actividad_economica_trimestral.xlsx",
    "subtitle": "BCV - actividades económicas",
  },
}


def now_utc() -> str:
  return dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def clean_text(value: object) -> str:
  if value is None:
    return ""
  if isinstance(value, float) and value.is_integer():
    value = int(value)
  return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()


def parse_year(value: object) -> int | None:
  text = clean_text(value)
  match = re.search(r"(19\d{2}|20\d{2})", text)
  return int(match.group(1)) if match else None


def parse_quarter(value: object) -> int | None:
  text = clean_text(value).upper()
  for token, quarter in QUARTERS.items():
    if re.search(rf"\b{token}\s*(TRIM|TRIMESTRE)?\b", text):
      return quarter
  return None


def number(value: object) -> float | int | None:
  if value in ("", None):
    return None
  if isinstance(value, (int, float)):
    if isinstance(value, float) and value.is_integer():
      return int(value)
    return value
  try:
    parsed = float(str(value).replace(",", "."))
  except ValueError:
    return None
  return int(parsed) if parsed.is_integer() else parsed


def source_url(filename: str) -> str:
  return f"{BASE_URL}{filename}"


def row(
  dataset_id: str,
  table: str,
  source_file: str,
  base: str,
  price_type: str,
  measure: str,
  frequency: str,
  period: str,
  year: int | None,
  quarter: int | None,
  classification: str,
  component: str,
  value: object,
  unit: str,
  fetched_at: str,
  status: str = "",
) -> dict:
  meta = DATASETS[dataset_id]
  return {
    "dataset_id": dataset_id,
    "title": meta["title"],
    "table": table,
    "source_file": source_file,
    "base": base,
    "price_type": price_type,
    "measure": measure,
    "frequency": frequency,
    "period": period,
    "year": year,
    "quarter": quarter,
    "classification": classification,
    "component": component,
    "value": number(value),
    "unit": unit,
    "status": status,
    "source": SOURCE,
    "source_url": source_url(source_file),
    "fetched_at": fetched_at,
  }


def xls_rows(path: Path, sheet_name: str) -> list[list[object]]:
  book = xlrd.open_workbook(str(path))
  sheet = book.sheet_by_name(sheet_name)
  return [sheet.row_values(idx) for idx in range(sheet.nrows)]


def parse_horizontal_table(
  dataset_id: str,
  path: Path,
  sheet_name: str,
  header_row: int,
  start_row: int,
  table: str,
  base: str,
  price_type: str,
  measure: str,
  unit: str,
  fetched_at: str,
  component_col: int = 0,
  code_col: int | None = None,
  first_value_col: int | None = None,
) -> list[dict]:
  rows = xls_rows(path, sheet_name)
  headers = rows[header_row]
  output = []
  current_classification = ""
  for raw in rows[start_row:]:
    component = clean_text(raw[component_col] if component_col < len(raw) else "")
    if not component:
      continue
    if all(number(cell) is None for cell in raw[component_col + 1:]):
      current_classification = component
      continue
    if code_col is not None:
      component = clean_text(raw[code_col]) or component
    value_start = first_value_col if first_value_col is not None else component_col + 1
    for col_idx in range(value_start, len(headers)):
      year = parse_year(headers[col_idx] if col_idx < len(headers) else "")
      value = raw[col_idx] if col_idx < len(raw) else None
      if year is None or number(value) is None:
        continue
      output.append(row(
        dataset_id,
        table,
        path.name,
        base,
        price_type,
        measure,
        "Anual",
        str(year),
        year,
        None,
        current_classification,
        component,
        value,
        unit,
        fetched_at,
        "Preliminar" if "(*)" in clean_text(headers[col_idx]) else "",
      ))
  return output


def parse_sector_xls(path: Path, sheet_name: str, dataset_id: str, measure: str, table: str, fetched_at: str, quarterly: bool) -> list[dict]:
  rows = xls_rows(path, sheet_name)
  components = ["Total", "Sector público", "Sector privado", "Impuestos netos sobre los productos"]
  output = []
  current_year = None
  for raw in rows:
    first = clean_text(raw[1] if len(raw) > 1 else "")
    year = parse_year(first)
    if year:
      current_year = year
      if not quarterly and len(raw) >= 5:
        for idx, component in enumerate(components, start=2):
          value = raw[idx] if idx < len(raw) else None
          if number(value) is not None:
            output.append(row(dataset_id, table, path.name, "1997", "Precios constantes", measure, "Anual", str(year), year, None, "Sector institucional", component, value, "Bolívares" if measure == "Nivel" else "Porcentaje", fetched_at, "Preliminar" if "(*)" in first else ""))
      continue
    if quarterly and current_year:
      quarter = parse_quarter(first)
      if quarter:
        period = f"{current_year}-T{quarter}"
        for idx, component in enumerate(components, start=2):
          value = raw[idx] if idx < len(raw) else None
          if number(value) is not None:
            output.append(row(dataset_id, table, path.name, "1997", "Precios constantes", measure, "Trimestral", period, current_year, quarter, "Sector institucional", component, value, "Bolívares" if measure == "Nivel" else "Porcentaje", fetched_at, "Preliminar"))
  return output


def parse_sector_xlsx(path: Path, dataset_id: str, fetched_at: str, quarterly: bool) -> list[dict]:
  wb = load_workbook(path, read_only=True, data_only=True)
  ws = wb.active
  components = ["Total", "Sector público", "Sector privado", "Impuestos netos sobre los productos"]
  output = []
  current_year = None
  for raw in ws.iter_rows(values_only=True):
    first = clean_text(raw[1] if len(raw) > 1 else "")
    year = parse_year(first)
    if year:
      current_year = year
      if not quarterly:
        for idx, component in enumerate(components, start=2):
          value = raw[idx] if idx < len(raw) else None
          if number(value) is not None:
            output.append(row(dataset_id, "Variaciones porcentuales", path.name, "2007", "Precios constantes", "Variación porcentual interanual", "Anual", str(year), year, None, "Sector institucional", component, value, "Porcentaje", fetched_at, "Preliminar" if "(*)" in first else ""))
      continue
    if quarterly and current_year:
      quarter = parse_quarter(first)
      if quarter:
        period = f"{current_year}-T{quarter}"
        for idx, component in enumerate(components, start=2):
          value = raw[idx] if idx < len(raw) else None
          if number(value) is not None:
            output.append(row(dataset_id, "Variaciones porcentuales", path.name, "2007", "Precios constantes", "Variación porcentual interanual", "Trimestral", period, current_year, quarter, "Sector institucional", component, value, "Porcentaje", fetched_at, "Preliminar"))
  return output


def parse_activity_xlsx(path: Path, dataset_id: str, fetched_at: str, quarterly: bool) -> list[dict]:
  wb = load_workbook(path, read_only=True, data_only=True)
  ws = wb.active
  rows = list(ws.iter_rows(values_only=True))
  output = []
  if quarterly:
    year_headers = rows[6]
    quarter_headers = rows[7]
    years = []
    current_year = None
    for col_idx, header in enumerate(year_headers):
      year = parse_year(header)
      if year:
        current_year = year
      years.append(current_year)
    for raw in rows[9:]:
      component = clean_text(raw[1] if len(raw) > 1 else "")
      if not component or component.lower().startswith(("fuente", "(*)")):
        continue
      for col_idx in range(2, len(raw)):
        year = years[col_idx] if col_idx < len(years) else None
        quarter = parse_quarter(quarter_headers[col_idx] if col_idx < len(quarter_headers) else "")
        value = raw[col_idx]
        if year and quarter and number(value) is not None:
          output.append(row(dataset_id, "Variaciones porcentuales", path.name, "2007", "Precios constantes", "Variación porcentual interanual", "Trimestral", f"{year}-T{quarter}", year, quarter, "Actividad económica", component, value, "Porcentaje", fetched_at, "Preliminar"))
  else:
    headers = rows[6]
    for raw in rows[8:]:
      component = clean_text(raw[1] if len(raw) > 1 else "")
      if not component or component.lower().startswith(("fuente", "(*)")):
        continue
      for col_idx in range(2, len(raw)):
        year = parse_year(headers[col_idx] if col_idx < len(headers) else "")
        value = raw[col_idx]
        if year and number(value) is not None:
          output.append(row(dataset_id, "Variaciones porcentuales", path.name, "2007", "Precios constantes", "Variación porcentual interanual", "Anual", str(year), year, None, "Actividad económica", component, value, "Porcentaje", fetched_at, "Preliminar" if "(*)" in clean_text(headers[col_idx]) else ""))
  return output


def parse_activity_xls_quarterly(path: Path, sheet_name: str, dataset_id: str, measure: str, unit: str, fetched_at: str) -> list[dict]:
  rows = xls_rows(path, sheet_name)
  year_headers = rows[6]
  quarter_headers = rows[7]
  years = []
  current_year = None
  for header in year_headers:
    year = parse_year(header)
    if year:
      current_year = year
    years.append(current_year)
  output = []
  for raw in rows[8:]:
    component = clean_text(raw[1] if len(raw) > 1 else "")
    if not component or component.lower().startswith(("fuente", "(*)", "nota", "1/", "2/")):
      continue
    for col_idx in range(2, len(raw)):
      year = years[col_idx] if col_idx < len(years) else None
      quarter = parse_quarter(quarter_headers[col_idx] if col_idx < len(quarter_headers) else "")
      value = raw[col_idx]
      if year and quarter and number(value) is not None:
        output.append(row(
          dataset_id,
          "PIB por actividad económica",
          path.name,
          "1997",
          "Precios constantes",
          measure,
          "Trimestral",
          f"{year}-T{quarter}",
          year,
          quarter,
          "Actividad económica",
          component,
          value,
          unit,
          fetched_at,
          "Preliminar",
        ))
  return output


def parse_historical(fetched_at: str) -> list[dict]:
  path = RAW_DIR / "7_1_14_anual.xls"
  rows = xls_rows(path, "7_1_14")
  output = []
  for raw in rows:
    year = parse_year(raw[2] if len(raw) > 2 else "")
    if year is None:
      continue
    constant = number(raw[4] if len(raw) > 4 else None)
    current = number(raw[6] if len(raw) > 6 else None)
    if constant is not None:
      output.append(row("bcv_pib_historico_anual", "PIB histórico anual", path.name, "1957/1968/1984/1997", "Precios constantes", "Nivel", "Anual", str(year), year, None, "Total economía", "Producto interno bruto", constant, "Millones de bolívares a precios constantes", fetched_at))
    if current is not None:
      output.append(row("bcv_pib_historico_anual", "PIB histórico anual", path.name, "Escala monetaria vigente desde 2018", "Precios corrientes", "Nivel", "Anual", str(year), year, None, "Total economía", "Producto interno bruto", current, "Millones de bolívares corrientes", fetched_at))
  return output


def collect_rows(fetched_at: str) -> dict[str, list[dict]]:
  datasets = {key: [] for key in DATASETS}
  datasets["bcv_pib_historico_anual"].extend(parse_historical(fetched_at))

  path = RAW_DIR / "7_1_7_anual.xls"
  datasets["bcv_pib_demanda_anual"].extend(parse_horizontal_table("bcv_pib_demanda_anual", path, "Precios Constantes", 5, 6, "Componentes del PIB", "1997", "Precios constantes", "Nivel", "Bolívares", fetched_at, component_col=2, code_col=2, first_value_col=3))
  datasets["bcv_pib_demanda_anual"].extend(parse_horizontal_table("bcv_pib_demanda_anual", path, "Precios Corrientes", 5, 6, "Componentes del PIB", "Escala monetaria vigente desde 2018", "Precios corrientes", "Nivel", "Bolívares", fetched_at, component_col=2, code_col=2, first_value_col=3))

  datasets["bcv_pib_sector_institucional_anual"].extend(parse_sector_xls(RAW_DIR / "5_2_1_anual.xls", "PIB", "bcv_pib_sector_institucional_anual", "Nivel", "PIB por sector institucional", fetched_at, quarterly=False))
  datasets["bcv_pib_sector_institucional_anual"].extend(parse_sector_xls(RAW_DIR / "5_2_1_anual.xls", "Var %", "bcv_pib_sector_institucional_anual", "Variación porcentual interanual", "PIB por sector institucional", fetched_at, quarterly=False))
  datasets["bcv_pib_sector_institucional_anual"].extend(parse_sector_xlsx(RAW_DIR / "5_2_1_si_anual.xlsx", "bcv_pib_sector_institucional_anual", fetched_at, quarterly=False))

  datasets["bcv_pib_sector_institucional_trimestral"].extend(parse_sector_xls(RAW_DIR / "5_2_1_trim.xls", "PIB", "bcv_pib_sector_institucional_trimestral", "Nivel", "PIB por sector institucional", fetched_at, quarterly=True))
  datasets["bcv_pib_sector_institucional_trimestral"].extend(parse_sector_xls(RAW_DIR / "5_2_1_trim.xls", "Var %", "bcv_pib_sector_institucional_trimestral", "Variación porcentual interanual", "PIB por sector institucional", fetched_at, quarterly=True))
  datasets["bcv_pib_sector_institucional_trimestral"].extend(parse_sector_xlsx(RAW_DIR / "5_2_1_si_trim.xlsx", "bcv_pib_sector_institucional_trimestral", fetched_at, quarterly=True))

  datasets["bcv_pib_actividad_economica_anual"].extend(parse_horizontal_table("bcv_pib_actividad_economica_anual", RAW_DIR / "5_2_4_anual.xls", "PIB Niveles", 6, 8, "PIB por actividad económica", "1997", "Precios constantes", "Nivel", "Bolívares", fetched_at, component_col=1, first_value_col=2))
  datasets["bcv_pib_actividad_economica_anual"].extend(parse_horizontal_table("bcv_pib_actividad_economica_anual", RAW_DIR / "5_2_4_anual.xls", "V% Porcentual", 6, 8, "PIB por actividad económica", "1997", "Precios constantes", "Variación porcentual interanual", "Porcentaje", fetched_at, component_col=1, first_value_col=2))
  datasets["bcv_pib_actividad_economica_anual"].extend(parse_activity_xlsx(RAW_DIR / "5_2_4_ae_anual.xlsx", "bcv_pib_actividad_economica_anual", fetched_at, quarterly=False))

  datasets["bcv_pib_actividad_economica_trimestral"].extend(parse_activity_xls_quarterly(RAW_DIR / "5_2_4_trim.xls", "PIB Niveles", "bcv_pib_actividad_economica_trimestral", "Nivel", "Bolívares", fetched_at))
  datasets["bcv_pib_actividad_economica_trimestral"].extend(parse_activity_xls_quarterly(RAW_DIR / "5_2_4_trim.xls", "PIB V% Puntual", "bcv_pib_actividad_economica_trimestral", "Variación porcentual interanual", "Porcentaje", fetched_at))
  datasets["bcv_pib_actividad_economica_trimestral"].extend(parse_activity_xlsx(RAW_DIR / "5_2_4_ae_trim.xlsx", "bcv_pib_actividad_economica_trimestral", fetched_at, quarterly=True))

  for rows in datasets.values():
    rows.sort(key=lambda item: (item["dataset_id"], item["component"], item["year"] or 0, item["quarter"] or 0, item["measure"], item["price_type"]))
  return datasets


def write_outputs(dataset_id: str, rows: list[dict], fetched_at: str) -> dict:
  meta = DATASETS[dataset_id]
  JSON_DIR.mkdir(parents=True, exist_ok=True)
  CSV_DIR.mkdir(parents=True, exist_ok=True)
  EXCEL_DIR.mkdir(parents=True, exist_ok=True)

  json_path = JSON_DIR / f"ove_{dataset_id}.json"
  csv_path = CSV_DIR / f"ove_{dataset_id}.csv"
  excel_path = EXCEL_DIR / meta["filename"]
  first_year = min((item["year"] for item in rows if item["year"]), default=None)
  last_period = max((item["period"] for item in rows if item["period"]), default=None)
  latest = max(rows, key=lambda item: (item["year"] or 0, item["quarter"] or 0)) if rows else None
  metadata = {
    "dataset_id": dataset_id,
    "title": meta["title"],
    "source": SOURCE,
    "source_url": SOURCE_PAGE,
    "frequency": "annual/quarterly" if any(item["frequency"] == "Trimestral" for item in rows) else "annual",
    "status": "official_source_normalized",
    "last_fetched_at": fetched_at,
    "first_year": first_year,
    "last_period": last_period,
    "records": len(rows),
    "excel": str(excel_path.relative_to(ROOT)),
    "csv": str(csv_path.relative_to(ROOT)),
    "json": str(json_path.relative_to(ROOT)),
    "latest": latest,
    "notes": "Normalizado desde workbooks oficiales enlazados por la página Producto Interno Bruto del Banco Central de Venezuela.",
  }

  json_path.write_text(json.dumps({"metadata": metadata, "observations": rows}, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
  with csv_path.open("w", newline="", encoding="utf-8") as handle:
    writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
    writer.writeheader()
    writer.writerows(rows)

  wb = Workbook()
  ws = wb.active
  ws.title = "datos"
  format_sheet(ws, meta["title"], meta["subtitle"])
  write_table(ws, FIELDS, rows)
  meta_ws = wb.create_sheet("metadatos")
  format_sheet(meta_ws, "Metadatos", meta["title"])
  write_key_values(meta_ws, [(key, json.dumps(value, ensure_ascii=False) if isinstance(value, (dict, list)) else value) for key, value in metadata.items()])
  wb.save(excel_path)
  return metadata


def update_catalog(entries: list[dict], fetched_at: str) -> None:
  catalog_path = CATALOG_DIR / "ove_bcv_producto_interno_bruto_operaciones.json"
  payload = {
    "metadata": {
      "dataset_id": "bcv_producto_interno_bruto_operaciones",
      "title": "Producto Interno Bruto - operaciones normalizadas para OVE",
      "source": SOURCE,
      "source_url": SOURCE_PAGE,
      "status": "official_source_normalized",
      "last_fetched_at": fetched_at,
      "datasets": len(entries),
      "notes": "Catálogo de Excel OVE creados a partir de workbooks oficiales del BCV para el apartado PIB.",
    },
    "datasets": entries,
  }
  CATALOG_DIR.mkdir(parents=True, exist_ok=True)
  catalog_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def main() -> int:
  fetched_at = now_utc()
  datasets = collect_rows(fetched_at)
  entries = [write_outputs(dataset_id, rows, fetched_at) for dataset_id, rows in datasets.items()]
  update_catalog(entries, fetched_at)
  print(json.dumps({"dataset": "BCV PIB operation workbooks", "generated_at": fetched_at, "outputs": [entry["excel"] for entry in entries]}, ensure_ascii=False))
  return 0


if __name__ == "__main__":
  raise SystemExit(main())
